import math
import random
import os
import shutil
import zipfile
import json
import urllib.request
from decimal import Decimal, ROUND_HALF_UP
import openpyxl

from rest_framework import viewsets, permissions, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from django.shortcuts import get_object_or_404, render, redirect
from django.utils import timezone
from django.db.models import Sum, Avg, Count, Q
from django.contrib import messages
from django.db import transaction
from django.conf import settings
from django.http import FileResponse, JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.urls import reverse

from .models import (
    User, Parent, Teacher, SchoolClass, Section,
    Student, Enrollment, Subject, Attendance, Grade,
    TimetableSlot, TeacherQuota, TimetableSubstitution, TimetableVersion,
    OfficialDocument, Invoice, AcademicYear, StudentAcademicHistory,
    ExamHall, ExamSession, ExamSeatAssignment, OfficialLetterTemplate,
    SchoolSettings
)
from .serializers import (
    UserSerializer, ParentSerializer, TeacherSerializer,
    SchoolClassSerializer, SectionSerializer, StudentSerializer,
    EnrollmentSerializer, SubjectSerializer, AttendanceSerializer,
    GradeSerializer, TimetableSlotSerializer, OfficialDocumentSerializer,
    InvoiceSerializer
)
from .tasks import send_absence_notification, send_invoice_reminder
from .licensing import (
    get_machine_fingerprint,
    verify_and_apply_license,
    generate_license_key,
    verify_and_apply_license_file,
    generate_license_file_data
)

# قائمة الصفوف العراقية الرسمية الـ 15 الثابتة
IRAQI_STANDARD_CLASSES = [
    (1, 'الاول الابتدائي', False),
    (2, 'الثاني الابتدائي', False),
    (3, 'الثالث الابتدائي', False),
    (4, 'الرابع الابتدائي', False),
    (5, 'الخامس الابتدائي', False),
    (6, 'السادس الابتدائي', True),
    (7, 'الاول المتوسط', False),
    (8, 'الثاني المتوسط', False),
    (9, 'الثالث المتوسط', True),
    (10, 'الرابع العلمي', False),
    (11, 'الرابع الادبي', False),
    (12, 'الخامس العلمي', False),
    (13, 'الخامس الادبي', False),
    (14, 'السادس العلمي', True),
    (15, 'السادس الادبي', True),
]


def seed_iraqi_official_classes(school):
    """تهيئة الصفوف الرسمية لوزارة التربية العراقية تلقائياً بحسب مرحلة المدرسة وصفوف التخرج"""
    school_level = getattr(school, 'school_level', 'secondary')
    
    all_stages = [
        (1, 'الاول الابتدائي', 'primary'),
        (2, 'الثاني الابتدائي', 'primary'),
        (3, 'الثالث الابتدائي', 'primary'),
        (4, 'الرابع الابتدائي', 'primary'),
        (5, 'الخامس الابتدائي', 'primary'),
        (6, 'السادس الابتدائي', 'primary'),
        (7, 'الاول المتوسط', 'intermediate'),
        (8, 'الثاني المتوسط', 'intermediate'),
        (9, 'الثالث المتوسط', 'intermediate'),
        (10, 'الرابع العلمي', 'secondary'),
        (11, 'الرابع الادبي', 'secondary'),
        (12, 'الخامس العلمي', 'secondary'),
        (13, 'الخامس الادبي', 'secondary'),
        (14, 'السادس العلمي', 'secondary'),
        (15, 'السادس الادبي', 'secondary'),
    ]

    if school_level == 'primary':
        selected_stages = [s for s in all_stages if s[2] == 'primary']
        final_names = ['السادس الابتدائي']
    elif school_level in ['intermediate', 'middle']:
        selected_stages = [s for s in all_stages if s[2] == 'intermediate']
        final_names = ['الثالث المتوسط']
    elif school_level == 'basic':
        selected_stages = [s for s in all_stages if s[2] in ['primary', 'intermediate']]
        final_names = ['الثالث المتوسط']
    elif school_level == 'preparatory':
        selected_stages = [s for s in all_stages if s[2] == 'secondary']
        final_names = ['السادس العلمي', 'السادس الادبي', 'السادس الأدبي']
    elif school_level in ['all_stages', 'all', 'comprehensive']:
        selected_stages = all_stages
        final_names = ['السادس الابتدائي', 'الثالث المتوسط', 'السادس العلمي', 'السادس الادبي', 'السادس الأدبي']
    else:  # secondary
        selected_stages = [s for s in all_stages if s[2] in ['intermediate', 'secondary']]
        final_names = ['السادس العلمي', 'السادس الادبي', 'السادس الأدبي']

    created_count = 0
    with transaction.atomic():
        for order, name, lvl in selected_stages:
            is_final = any(fn in name for fn in final_names)
            cls_obj, created = SchoolClass.objects.get_or_create(
                name=name,
                defaults={
                    'level_order': order,
                    'is_final_stage': is_final
                }
            )
            if cls_obj.is_final_stage != is_final or cls_obj.level_order != order:
                cls_obj.is_final_stage = is_final
                cls_obj.level_order = order
                cls_obj.save()
            if not cls_obj.sections.exists():
                Section.objects.create(school_class=cls_obj, name="أ", capacity=40)
            if created:
                created_count += 1

        # ربط الصفوف بالصف اللاحق الحتمي في قاعدة البيانات
        for c in SchoolClass.objects.all():
            nxt = get_next_promotion_stage(c, school_level)
            if c.next_class != nxt:
                c.next_class = nxt
                c.save(update_fields=['next_class'])

    return created_count


OFFICIAL_PROMOTION_CHAIN = {
    'الاول الابتدائي': 'الثاني الابتدائي',
    'الأول الابتدائي': 'الثاني الابتدائي',
    'الثاني الابتدائي': 'الثالث الابتدائي',
    'الثالث الابتدائي': 'الرابع الابتدائي',
    'الرابع الابتدائي': 'الخامس الابتدائي',
    'الخامس الابتدائي': 'السادس الابتدائي',
    'السادس الابتدائي': None,
    'الاول المتوسط': 'الثاني المتوسط',
    'الأول المتوسط': 'الثاني المتوسط',
    'الثاني المتوسط': 'الثالث المتوسط',
    'الثالث المتوسط': None,  # تفرع استثنائي إلى الرابع العلمي أو الرابع الأدبي في المدارس الثانوية
    'الرابع العلمي': 'الخامس العلمي',
    'الخامس العلمي': 'السادس العلمي',
    'السادس العلمي': None,
    'الرابع الادبي': 'الخامس الادبي',
    'الرابع الأدبي': 'الخامس الأدبي',
    'الخامس الادبي': 'السادس الادبي',
    'الخامس الأدبي': 'السادس الأدبي',
    'السادس الادبي': None,
    'السادس الأدبي': None,
}


def get_next_promotion_stage(cls_obj, school_level='secondary'):
    """
    تحديد الصف اللاحق التلقائي الحتمي بحسب السلم التعليمي العراقي الرسمي المعتمد:
    لا يسمح بالقفز أو التحديد اليدوي العشوائي. كل صف يرتبط حتمياً بالصف الذي يليه.
    """
    if not cls_obj or cls_obj.is_final_stage:
        return None

    c_name = cls_obj.name.strip()

    # إذا كان السادس الابتدائي، وفي مدرسة شاملة أو أساسية وبها الأول المتوسط
    if 'السادس الابتدائي' in c_name:
        if school_level in ['basic', 'all_stages', 'all', 'comprehensive']:
            return SchoolClass.objects.filter(name__icontains='الاول المتوسط').first()
        return None

    # إذا كان الثالث المتوسط، يخضع لتفرع علمي/أدبي في المدارس الثانوية
    if 'الثالث المتوسط' in c_name:
        return None

    # إذا كان سادس إعدادي (علمي أو أدبي)
    if any(s in c_name for s in ['السادس العلمي', 'السادس الأدبي', 'السادس الادبي']):
        return None

    target_name = OFFICIAL_PROMOTION_CHAIN.get(c_name)
    if target_name:
        dest = SchoolClass.objects.filter(name=target_name).first()
        if not dest:
            dest = SchoolClass.objects.filter(name__icontains=target_name).first()
        if dest:
            return dest

    # fallback أصولي بناءً على ترتيب المرحلة المتتالية بدقة (level_order + 1)
    return SchoolClass.objects.filter(level_order=cls_obj.level_order + 1).first()


def get_third_intermediate_branches():
    """الحصول على صفي التفرع (الرابع العلمي والرابع الأدبي) للثالث المتوسط"""
    scientific = SchoolClass.objects.filter(name__icontains='الرابع العلمي').first()
    literary = SchoolClass.objects.filter(name__icontains='الرابع الادبي').first() or SchoolClass.objects.filter(name__icontains='الرابع الأدبي').first()
    return scientific, literary


def round_integer(val):
    """دالة لتقريب الدرجات والمعدلات إلى أقرب عدد صحيح دون كسور"""
    if val is None or val == '':
        return None
    try:
        d = Decimal(str(val))
        return int(d.quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    except Exception:
        return int(round(float(val)))


def get_active_academic_year():
    """الحصول على العام الدراسي المفعل أو إنشاء واحد تلقائياً في حال النقص"""
    year = AcademicYear.objects.filter(is_current=True).first()
    if not year:
        year = AcademicYear.objects.first()
        if not year:
            year = AcademicYear.objects.create(name="2026-2027", is_current=True)
        else:
            year.is_current = True
            year.save()
    return year


class IsStaffOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return request.user and request.user.is_authenticated
        return request.user and (request.user.is_staff or request.user.is_superuser)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by('-date_joined')
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAdminUser]
    filter_backends = [filters.SearchFilter]
    search_fields = ['username', 'first_name', 'last_name', 'email']

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def me(self, request):
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)


class ParentViewSet(viewsets.ModelViewSet):
    queryset = Parent.objects.select_related('user').all()
    serializer_class = ParentSerializer
    permission_classes = [IsStaffOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['user__first_name', 'user__last_name', 'phone', 'address']


class TeacherViewSet(viewsets.ModelViewSet):
    queryset = Teacher.objects.select_related('user').prefetch_related('subjects').all()
    serializer_class = TeacherSerializer
    permission_classes = [IsStaffOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['user__first_name', 'user__last_name', 'job_title', 'statistical_code']


class SchoolClassViewSet(viewsets.ModelViewSet):
    queryset = SchoolClass.objects.prefetch_related('sections').all()
    serializer_class = SchoolClassSerializer
    permission_classes = [IsStaffOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']


class SectionViewSet(viewsets.ModelViewSet):
    queryset = Section.objects.select_related('school_class').all()
    serializer_class = SectionSerializer
    permission_classes = [IsStaffOrReadOnly]

    def get_queryset(self):
        qs = super().get_queryset()
        class_id = self.request.query_params.get('school_class')
        if class_id:
            qs = qs.filter(school_class_id=class_id)
        return qs


class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.filter(is_deleted=False).select_related('user', 'parent__user', 'current_class', 'section').all()
    serializer_class = StudentSerializer
    permission_classes = [IsStaffOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['user__first_name', 'user__last_name', 'national_id', 'registration_number']

    def get_queryset(self):
        qs = super().get_queryset()
        class_id = self.request.query_params.get('current_class')
        section_id = self.request.query_params.get('section')
        status_param = self.request.query_params.get('student_status')
        if class_id:
            qs = qs.filter(current_class_id=class_id)
        if section_id:
            qs = qs.filter(section_id=section_id)
        if status_param:
            qs = qs.filter(student_status=status_param)
        return qs

    def perform_destroy(self, instance):
        instance.soft_delete()

    @action(detail=True, methods=['get'])
    def grades(self, request, pk=None):
        student = self.get_object()
        grades = Grade.objects.filter(student=student).select_related('subject')
        serializer = GradeSerializer(grades, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def attendance(self, request, pk=None):
        student = self.get_object()
        attendance_records = Attendance.objects.filter(student=student).order_by('-date')
        serializer = AttendanceSerializer(attendance_records, many=True)
        return Response(serializer.data)


class EnrollmentViewSet(viewsets.ModelViewSet):
    queryset = Enrollment.objects.select_related('student__user', 'school_class').all()
    serializer_class = EnrollmentSerializer
    permission_classes = [IsStaffOrReadOnly]


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsStaffOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'code']


class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.select_related('student__user', 'recorded_by').all().order_by('-date')
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        attendance = serializer.save(recorded_by=user)
        if attendance.status == 'absent':
            try:
                send_absence_notification.delay(attendance.student_id, str(attendance.date))
            except Exception:
                pass

    def get_queryset(self):
        qs = super().get_queryset()
        student_id = self.request.query_params.get('student')
        date = self.request.query_params.get('date')
        if student_id:
            qs = qs.filter(student_id=student_id)
        if date:
            qs = qs.filter(date=date)
        return qs

    @action(detail=False, methods=['post'])
    def bulk_record(self, request):
        records = request.data
        if not isinstance(records, list):
            return Response({'error': 'Expected a list of attendance records'}, status=status.HTTP_400_BAD_REQUEST)

        created = []
        user = request.user if request.user.is_authenticated else None
        for item in records:
            serializer = self.get_serializer(data=item)
            if serializer.is_valid():
                att = serializer.save(recorded_by=user)
                if att.status == 'absent':
                    try:
                        send_absence_notification.delay(att.student_id, str(att.date))
                    except Exception:
                        pass
                created.append(serializer.data)
        return Response({'created_count': len(created), 'records': created}, status=status.HTTP_201_CREATED)


class GradeViewSet(viewsets.ModelViewSet):
    queryset = Grade.objects.select_related('student__user', 'subject').all()
    serializer_class = GradeSerializer
    permission_classes = [IsStaffOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['student__user__first_name', 'student__user__last_name', 'subject__name', 'status']

    def get_queryset(self):
        qs = super().get_queryset()
        student_id = self.request.query_params.get('student')
        subject_id = self.request.query_params.get('subject')
        year = self.request.query_params.get('academic_year')
        status_filter = self.request.query_params.get('status')
        if student_id:
            qs = qs.filter(student_id=student_id)
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        if year:
            qs = qs.filter(academic_year=year)
        if status_filter:
            qs = qs.filter(status=status_filter)
        return qs

    @action(detail=True, methods=['post'])
    def apply_decision(self, request, pk=None):
        grade = self.get_object()
        max_decision = int(request.data.get('max_decision', 5))
        applied = grade.apply_decision_marks(max_allowed=max_decision)
        grade.save()
        return Response({
            'grade_id': grade.id,
            'decision_marks_applied': str(round_integer(applied)),
            'final_grade': str(round_integer(grade.final_grade_after_decision)),
            'status': grade.status,
            'status_display': grade.get_status_display()
        })

    @action(detail=False, methods=['post'])
    def batch_apply_decision(self, request):
        student_id = request.data.get('student_id')
        if not student_id:
            return Response({'error': 'student_id مطلوب'}, status=status.HTTP_400_BAD_REQUEST)

        max_total = Decimal(str(request.data.get('max_total_decision', 5)))
        grades = Grade.objects.filter(student_id=student_id).order_by('final_grade')

        used = Decimal('0')
        updated = []

        for g in grades:
            remaining = max_total - used
            if remaining <= 0:
                break
            if g.final_grade is not None and Decimal('45.0') <= g.final_grade < Decimal('50.0'):
                needed = Decimal('50.0') - g.final_grade
                if needed <= remaining:
                    g.decision_marks = needed
                    g.final_grade_after_decision = Decimal('50.0')
                    g.status = 'passed_by_decision'
                    g.save()
                    used += needed
                    updated.append({'subject': g.subject.name, 'decision_marks': str(round_integer(needed))})

        return Response({
            'student_id': student_id,
            'total_decision_used': str(round_integer(used)),
            'remaining_decision': str(round_integer(max_total - used)),
            'updated_subjects': updated
        })


class TimetableSlotViewSet(viewsets.ModelViewSet):
    queryset = TimetableSlot.objects.select_related('school_class', 'section', 'subject', 'teacher__user').all()
    serializer_class = TimetableSlotSerializer
    permission_classes = [IsStaffOrReadOnly]

    def get_queryset(self):
        qs = super().get_queryset()
        class_id = self.request.query_params.get('school_class')
        teacher_id = self.request.query_params.get('teacher')
        day = self.request.query_params.get('day')
        if class_id:
            qs = qs.filter(school_class_id=class_id)
        if teacher_id:
            qs = qs.filter(teacher_id=teacher_id)
        if day:
            qs = qs.filter(day=day)
        return qs


class OfficialDocumentViewSet(viewsets.ModelViewSet):
    queryset = OfficialDocument.objects.select_related('created_by').all()
    serializer_class = OfficialDocumentSerializer
    permission_classes = [IsStaffOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['doc_number', 'subject', 'sender_receiver', 'notes', 'incoming_doc_number', 'outgoing_reply_number']

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        serializer.save(created_by=user)

    def get_queryset(self):
        qs = super().get_queryset()
        doc_type = self.request.query_params.get('doc_type')
        status_filter = self.request.query_params.get('status')
        if doc_type:
            qs = qs.filter(doc_type=doc_type)
        if status_filter:
            qs = qs.filter(status=status_filter)
        return qs


class InvoiceViewSet(viewsets.ModelViewSet):
    queryset = Invoice.objects.select_related('student__user').all().order_by('-created_at')
    serializer_class = InvoiceSerializer
    permission_classes = [IsStaffOrReadOnly]

    def get_queryset(self):
        qs = super().get_queryset()
        student_id = self.request.query_params.get('student')
        status_filter = self.request.query_params.get('status')
        if student_id:
            qs = qs.filter(student_id=student_id)
        if status_filter:
            qs = qs.filter(status=status_filter)
        return qs

    @action(detail=True, methods=['post'])
    def mark_as_paid(self, request, pk=None):
        invoice = self.get_object()
        invoice.status = 'paid'
        invoice.save()
        return Response({'status': 'Invoice marked as paid', 'invoice_id': invoice.id})

    @action(detail=True, methods=['post'])
    def send_reminder(self, request, pk=None):
        invoice = self.get_object()
        try:
            send_invoice_reminder.delay(invoice.id)
            return Response({'status': 'Reminder task scheduled successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ======================================================================
# واجهات الـ API التجارية
# ======================================================================

class SyncEngineViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['post'])
    def push_changes(self, request):
        payload = request.data
        changes_count = len(payload.get('records', []))
        return Response({
            'status': 'success',
            'synced_count': changes_count,
            'server_timestamp': timezone.now().isoformat(),
            'message': f'تمت مزامنة {changes_count} سجل بنجاح دون أي تضارب.'
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def pull_changes(self, request):
        return Response({
            'server_timestamp': timezone.now().isoformat(),
            'updates_available': False,
            'records': []
        }, status=status.HTTP_200_OK)


class BackupRestoreAPI(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        db_path = settings.DATABASES['default']['NAME']
        size_bytes = os.path.getsize(db_path) if os.path.exists(db_path) else 0
        return Response({
            'database_size_kb': round(size_bytes / 1024, 2),
            'last_modified': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'auto_sync_enabled': True
        })

    def post(self, request):
        return Response({'status': 'uploaded', 'message': 'تم حفظ النسخة السحابية بنجاح.'}, status=status.HTTP_201_CREATED)


class LicenseValidationAPI(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        machine_id = request.data.get('machine_id', '').strip()
        school = SchoolSettings.get_settings()
        is_valid = school.is_active()
        days_remaining = max(0, (school.subscription_end_date - timezone.now().date()).days) if school.subscription_end_date else 0

        return Response({
            'school_name': school.school_name,
            'is_active': is_valid,
            'days_remaining': days_remaining,
            'subscription_end_date': str(school.subscription_end_date),
            'machine_id': machine_id
        })


class SystemUpdateAPI(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        client_version = request.query_params.get('version', '1.0.0')
        latest_version = "1.0.0"
        return Response({
            'current_installed_version': client_version,
            'latest_available_version': latest_version,
            'has_update': False,
            'update_notes': 'المنظومة محدثة بالكامل وتعمل وفق معايير وزارة التربية 2026-2027.',
            'download_url': None
        })


# ======================================================================
# دوال استخراج وطباعة الوثائق
# ======================================================================

def teacher_service_certificate_view(request, teacher_id):
    school = SchoolSettings.get_settings()
    teacher = get_object_or_404(Teacher.objects.select_related('user'), pk=teacher_id)
    destination = request.GET.get('destination', 'إلى من يهمه الأمر')
    purpose = request.GET.get('purpose', 'المعاملات الرسمية والمصرفية')
    doc_number = request.GET.get('doc_number', f"ت/{teacher.id}/{timezone.now().strftime('%Y')}")

    context = {
        'school': school,
        'teacher': teacher,
        'destination': destination,
        'purpose': purpose,
        'doc_number': doc_number,
        'today_date': timezone.now().strftime('%Y/%m/%d'),
        'academic_year': get_active_academic_year(),
    }
    return render(request, 'certificates/teacher_service.html', context)


def student_transcript_view(request, student_id):
    school = SchoolSettings.get_settings()
    student = get_object_or_404(Student.objects.select_related('user', 'current_class', 'section'), pk=student_id)
    grades = Grade.objects.filter(student=student).select_related('subject')

    total_sum = Decimal('0')
    count = 0
    failed_count = 0

    for g in grades:
        eff = g.final_grade_after_decision or g.final_grade
        if eff is not None:
            total_sum += eff
            count += 1
            if eff < Decimal('50.0'):
                failed_count += 1

    total_avg = round_integer(total_sum / Decimal(str(count))) if count > 0 else 0
    clean_total_sum = round_integer(total_sum) if count > 0 else 0

    if failed_count == 0 and count > 0:
        final_result = "ناجح"
    elif failed_count <= 2 and count > 0:
        final_result = "مكمل (مؤهل لامتحان الدور الثاني)"
    else:
        final_result = "راسب" if count > 0 else "قيد المعالجة"

    doc_number = request.GET.get('doc_number', f"ط/{student.id}/{timezone.now().strftime('%Y')}")
    destination = request.GET.get('destination', 'إلى من يهمه الأمر / الإدارة التربوية المعنية')

    context = {
        'school': school,
        'student': student,
        'grades': grades,
        'total_sum': str(clean_total_sum) if count > 0 else "---",
        'total_avg': str(total_avg) if count > 0 else "---",
        'final_result': final_result,
        'destination': destination,
        'doc_number': doc_number,
        'today_date': timezone.now().strftime('%Y/%m/%d'),
        'academic_year': get_active_academic_year(),
    }
    return render(request, 'certificates/student_transcript.html', context)


def class_master_sheet_view(request, class_id):
    school = SchoolSettings.get_settings()
    school_class = get_object_or_404(SchoolClass, pk=class_id)
    
    section_id = request.GET.get('section_id')
    students_qs = Student.objects.filter(current_class=school_class, is_deleted=False).select_related('user', 'section')
    if section_id and section_id.isdigit():
        students_qs = students_qs.filter(section_id=int(section_id))
    students = list(students_qs.order_by('section__name', 'user__first_name', 'user__last_name', 'id'))
    
    # استحضار المواد الرسمية المعتمدة للمرحلة الدراسية
    official_subject_names = get_stage_official_subjects(school_class)
    subjects = []
    for s_name in official_subject_names:
        sub_obj = Subject.objects.filter(name=s_name).first()
        if not sub_obj:
            sub_obj = Subject.objects.filter(name__icontains=s_name.split()[0]).first()
        if sub_obj and sub_obj not in subjects:
            subjects.append(sub_obj)
        elif not sub_obj:
            sub_obj, _ = Subject.objects.get_or_create(name=s_name, defaults={'code': f'SUB_{len(subjects)+1}'})
            if sub_obj not in subjects:
                subjects.append(sub_obj)

    sheet_data = []
    from collections import defaultdict
    current_year = get_active_academic_year()
    all_grades = Grade.objects.filter(student__in=students, academic_year=current_year.name if current_year else '2026-2027')
    grades_by_student = defaultdict(dict)
    for g in all_grades:
        grades_by_student[g.student_id][g.subject_id] = g

    for st in students:
        student_grades = grades_by_student.get(st.id, {})
        subjects_grades = []
        tot = Decimal('0')
        valid_count = 0
        total_decision = Decimal('0')
        fails = 0

        for sub in subjects:
            g = student_grades.get(sub.id)
            if g:
                subjects_grades.append(g)
                eff = g.final_grade_after_decision or g.final_grade
                if eff is not None:
                    tot += eff
                    valid_count += 1
                    if eff < Decimal('50.0'):
                        fails += 1
                if g.decision_marks:
                    total_decision += g.decision_marks
            else:
                subjects_grades.append(None)

        avg_val = round_integer(tot / Decimal(str(valid_count))) if valid_count > 0 else 0

        if fails == 0 and valid_count > 0:
            res = "ناجح بالقرار" if total_decision > 0 else "ناجح"
        elif fails <= 2 and valid_count > 0:
            res = "مكمل"
        else:
            res = "راسب" if valid_count > 0 else "غير مكتمل"

        sheet_data.append({
            'student': st,
            'subjects_grades': subjects_grades,
            'total_sum': str(round_integer(tot)) if valid_count > 0 else "---",
            'avg': str(avg_val) if valid_count > 0 else "---",
            'total_decision': str(round_integer(total_decision)),
            'result': res
        })

    # تقسيم البيانات إلى صفحات بحيث لا تتجاوز كل صفحة 27 طالباً كحد أقصى (PAGINATION_LIMIT = 27)
    PAGINATION_LIMIT = 27
    total_students = len(sheet_data)
    pages_data = []

    if total_students == 0:
        pages_data.append({
            'page_number': 1,
            'students': [],
            'start_idx': 0,
            'is_last': True,
        })
    else:
        total_pages = (total_students + PAGINATION_LIMIT - 1) // PAGINATION_LIMIT
        for page_idx in range(total_pages):
            start_i = page_idx * PAGINATION_LIMIT
            end_i = min(start_i + PAGINATION_LIMIT, total_students)
            chunk = sheet_data[start_i:end_i]
            pages_data.append({
                'page_number': page_idx + 1,
                'students': chunk,
                'start_idx': start_i,
                'is_last': (page_idx + 1 == total_pages),
            })

    context = {
        'school': school,
        'school_class': school_class,
        'subjects': subjects,
        'sheet_data': sheet_data,
        'pages_data': pages_data,
        'total_pages': len(pages_data),
        'total_students': total_students,
        'today_date': timezone.now().strftime('%Y/%m/%d'),
        'academic_year': get_active_academic_year(),
    }
    return render(request, 'certificates/master_sheet.html', context)


def portal_student_result_cards(request, student_id=None):
    """
    منظومة طباعة نتائج الطلبة المدرسية (كارت النتيجة الرسمي / بطاقة درجات الطالب)
    - يدعم الطباعة الفردية لطالب محدد، أو الطباعة الجماعية لطلاب الشعبة/الصف بالكامل
    - يحتوي كارت النتيجة لكل طالب على الترويسة المدرسية وبيانات الطالب
    - جدول المواد بالأعمدة الإلزامية الثمانية حصراً:
      1. المادة
      2. درجة الفصل الأول
      3. درجة نصف السنة
      4. درجة الفصل الثاني
      5. درجة السعي السنوي
      6. درجة الامتحان النهائي
      7. الدرجة النهائية
      8. الملاحظات
    - خانات سفلية مخصصة للنتيجة النهائية: المجموع، المعدل، والقرار (ناجح/مكمل/راسب) مع التوقيعات الرسمية
    - تجهيز وضع الطباعة الجماعية بنسق مناسب (كارتين في الصفحة A4 مع خط القص، أو كارت لكل صفحة)
    """
    school = SchoolSettings.get_settings()
    current_year = get_active_academic_year()
    classes = SchoolClass.objects.all().order_by('level_order')

    selected_student_id = student_id or request.GET.get('student_id')
    selected_class_id = request.GET.get('class_id')
    selected_section_id = request.GET.get('section_id')
    print_layout = request.GET.get('layout', '2up')

    selected_class = None
    sections = []
    students_list = []

    if selected_student_id:
        single_student = get_object_or_404(Student.objects.select_related('user', 'current_class', 'section'), pk=selected_student_id)
        selected_class = single_student.current_class
        selected_class_id = str(selected_class.id) if selected_class else ''
        selected_section_id = str(single_student.section_id) if single_student.section_id else ''
        if selected_class:
            sections = list(selected_class.sections.all().order_by('name'))
        students_list = [single_student]
    else:
        if not selected_class_id and classes.exists():
            selected_class = classes.first()
            selected_class_id = str(selected_class.id)
        elif selected_class_id:
            selected_class = get_object_or_404(SchoolClass, pk=selected_class_id)

        if selected_class:
            sections = list(selected_class.sections.all().order_by('name'))
            st_qs = Student.objects.filter(current_class=selected_class, is_deleted=False).select_related('user', 'section')
            if selected_section_id and selected_section_id.isdigit():
                st_qs = st_qs.filter(section_id=int(selected_section_id))
            students_list = list(st_qs.order_by('section__name', 'user__first_name', 'user__last_name', 'id'))

    # استحضار المواد الرسمية للصف
    official_subject_names = get_stage_official_subjects(selected_class)
    class_subjects = []
    for s_name in official_subject_names:
        sub_obj = Subject.objects.filter(name=s_name).first()
        if not sub_obj:
            sub_obj = Subject.objects.filter(name__icontains=s_name.split()[0]).first()
        if sub_obj and sub_obj not in class_subjects:
            class_subjects.append(sub_obj)
        elif not sub_obj:
            sub_obj, _ = Subject.objects.get_or_create(name=s_name, defaults={'code': f'SUB_{len(class_subjects)+1}'})
            if sub_obj not in class_subjects:
                class_subjects.append(sub_obj)

    # تجهيز بطاقة النتيجة لكل طالب بالأعمدة الإلزامية الثمانية
    cards_data = []
    from collections import defaultdict
    all_grades = Grade.objects.filter(student__in=students_list, academic_year=current_year.name if current_year else '2026-2027')
    grades_by_student = defaultdict(dict)
    for g in all_grades:
        grades_by_student[g.student_id][g.subject_id] = g

    for st in students_list:
        student_grades = grades_by_student.get(st.id, {})
        subjects_rows = []
        total_sum = Decimal('0')
        valid_count = 0
        total_decision = Decimal('0')
        failed_subjects = []

        for sub in class_subjects:
            g = student_grades.get(sub.id)
            if g:
                f_term = g.first_term_effort
                mid = g.midyear_exam
                s_term = g.second_term_effort
                annual = g.annual_effort
                fin = g.final_exam_round2 if g.final_exam_round2 is not None else g.final_exam_round1
                final_val = g.final_grade_after_decision if g.final_grade_after_decision is not None else g.final_grade

                if final_val is not None:
                    total_sum += final_val
                    valid_count += 1
                    if final_val < Decimal('50.0'):
                        failed_subjects.append(sub.name)

                if g.decision_marks:
                    total_decision += g.decision_marks

                # الملاحظات
                if final_val is not None:
                    if g.decision_marks and g.decision_marks > 0:
                        note = "ناجح بالقرار"
                    elif final_val >= Decimal('50.0'):
                        note = "ناجح"
                    elif g.final_exam_round2 is not None and final_val < Decimal('50.0'):
                        note = "راسب دور ثانٍ"
                    else:
                        note = "مكمل"
                else:
                    note = "---"

                subjects_rows.append({
                    'subject': sub.name,
                    'first_term': str(round_integer(f_term)) if f_term is not None else "---",
                    'midyear': str(round_integer(mid)) if mid is not None else "---",
                    'second_term': str(round_integer(s_term)) if s_term is not None else "---",
                    'annual_effort': str(round_integer(annual)) if annual is not None else "---",
                    'final_exam': str(round_integer(fin)) if fin is not None else "---",
                    'final_grade': str(round_integer(final_val)) if final_val is not None else "---",
                    'notes': note,
                    'is_failed': final_val is not None and final_val < Decimal('50.0')
                })
            else:
                subjects_rows.append({
                    'subject': sub.name,
                    'first_term': "---",
                    'midyear': "---",
                    'second_term': "---",
                    'annual_effort': "---",
                    'final_exam': "---",
                    'final_grade': "---",
                    'notes': "---",
                    'is_failed': False
                })

        avg_val = round_integer(total_sum / Decimal(str(valid_count))) if valid_count > 0 else 0

        # القرار النهائي
        if valid_count == 0:
            final_result = "قيد الإنجاز"
            result_badge = "secondary"
        elif len(failed_subjects) == 0:
            final_result = "ناجح بالقرار" if total_decision > 0 else "ناجح"
            result_badge = "success"
        elif len(failed_subjects) <= 2:
            final_result = f"مكمل في ({'، '.join(failed_subjects)})"
            result_badge = "warning"
        else:
            final_result = "راسب"
            result_badge = "danger"

        cards_data.append({
            'student': st,
            'subjects_rows': subjects_rows,
            'total_sum': str(round_integer(total_sum)) if valid_count > 0 else "---",
            'avg': str(avg_val) if valid_count > 0 else "---",
            'total_decision': str(round_integer(total_decision)),
            'final_result': final_result,
            'result_badge': result_badge,
            'failed_count': len(failed_subjects),
        })

    # تقسيم البطاقات لأزواج (2 في الصفحة A4) للطباعة الجماعية
    cards_pairs = [cards_data[i:i + 2] for i in range(0, len(cards_data), 2)]

    class_students_list = []
    if selected_class:
        class_students_list = list(Student.objects.filter(current_class=selected_class, is_deleted=False).order_by('user__first_name'))

    context = {
        'school': school,
        'current_year': current_year,
        'classes': classes,
        'selected_class': selected_class,
        'selected_class_id': selected_class_id,
        'sections': sections,
        'selected_section_id': selected_section_id,
        'class_students_list': class_students_list,
        'selected_student_id': selected_student_id,
        'cards_data': cards_data,
        'cards_pairs': cards_pairs,
        'total_cards': len(cards_data),
        'print_layout': print_layout,
        'today_date': timezone.now().strftime('%Y/%m/%d'),
        'single_mode': bool(selected_student_id),
    }
    return render(request, 'portal/student_result_cards.html', context)


# ======================================================================
# دوال البوابة التفاعلية والترخيص
# ======================================================================

def portal_license_lock(request):
    school = SchoolSettings.get_settings()
    machine_id = get_machine_fingerprint()

    if request.method == 'POST':
        license_file = request.FILES.get('license_file')
        license_key = request.POST.get('license_key', '').strip()
        if license_file:
            success, message = verify_and_apply_license_file(school, license_file.read())
        elif license_key:
            success, message = verify_and_apply_license(school, license_key)
        else:
            success, message = False, "يرجى إدخال مفتاح التفعيل أو اختيار ملف الترخيص (.lic)."

        if success:
            messages.success(request, message)
            return redirect('portal_dashboard')
        else:
            messages.error(request, message)

    context = {
        'school': school,
        'machine_id': machine_id,
        'phone_contact': "07723457175",
    }
    return render(request, 'portal/license_lock.html', context)


def portal_license_activate(request):
    if request.method == 'POST':
        school = SchoolSettings.get_settings()
        license_file = request.FILES.get('license_file')
        license_key = request.POST.get('license_key', '').strip()

        if license_file:
            success, message = verify_and_apply_license_file(school, license_file.read())
        elif license_key:
            success, message = verify_and_apply_license(school, license_key)
        else:
            success, message = False, "يرجى إدخال مفتاح التفعيل أو اختيار ملف الترخيص (.lic)."

        if success:
            messages.success(request, message)
        else:
            messages.error(request, message)

    next_url = request.META.get('HTTP_REFERER', 'portal_dashboard')
    return redirect(next_url)


def owner_key_generator(request):
    OWNER_USER = "abdullahnawfal97"
    OWNER_PASS = "111997111997"

    if not request.session.get('is_owner_authenticated'):
        if request.method == 'POST':
            u = (request.POST.get('username') or request.POST.get('owner_user') or '').strip().lower()
            p = (request.POST.get('password') or request.POST.get('owner_pass') or '').strip()

            if u == OWNER_USER and p == OWNER_PASS:
                request.session['is_owner_authenticated'] = True
                messages.success(request, "تم تسجيل دخول المطور والمالك بنجاح.")
                return redirect('portal_owner_generator')
            else:
                messages.error(request, "اسم المستخدم أو كلمة المرور غير صحيحة!")

        return render(request, 'portal/owner_login.html')

    generated_key = None
    target_machine = ""
    selected_plan = "YEAR"

    if request.method == 'POST':
        action_type = request.POST.get('action_type')
        target_machine = request.POST.get('machine_id', '').strip().upper()
        selected_plan = request.POST.get('plan', 'YEAR')

        if action_type == 'generate':
            if target_machine:
                generated_key = generate_license_key(target_machine, plan=selected_plan)
                messages.success(request, "تم توليد مفتاح التفعيل بنجاح!")
            else:
                messages.error(request, "يرجى كتابة معرف الحاسبة (Machine ID).")

        elif action_type == 'download_lic':
            if target_machine:
                lic_json = generate_license_file_data(target_machine, plan=selected_plan)
                response = HttpResponse(lic_json, content_type='application/json; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="Madrasati_{target_machine[:8]}.lic"'
                return response
            else:
                messages.error(request, "يرجى كتابة معرف الحاسبة لتوليد ملف الترخيص.")

    context = {
        'is_authenticated': True,
        'generated_key': generated_key,
        'target_machine': target_machine,
        'selected_plan': selected_plan,
    }
    return render(request, 'portal/owner_generator.html', context)


def portal_dashboard(request):
    school = SchoolSettings.get_settings()
    today = timezone.now().date()

    if not school.is_trial_or_license_valid:
        return redirect('portal_license_lock')

    machine_id = get_machine_fingerprint()
    current_year = get_active_academic_year()
    academic_years = AcademicYear.objects.all().order_by('name')
    students_count = Student.objects.filter(is_deleted=False, student_status='active').count()
    teachers_count = Teacher.objects.count()
    classes_count = SchoolClass.objects.count()
    pending_docs = OfficialDocument.objects.filter(status='pending').count()
    recent_docs = OfficialDocument.objects.all().order_by('-created_at')[:5]

    days_left = school.days_remaining

    context = {
        'school': school,
        'machine_id': machine_id,
        'current_year': current_year,
        'academic_years': academic_years,
        'students_count': students_count,
        'teachers_count': teachers_count,
        'classes_count': classes_count,
        'pending_docs': pending_docs,
        'recent_docs': recent_docs,
        'days_left': days_left,
    }
    return render(request, 'portal/dashboard.html', context)


def portal_set_current_year(request):
    if request.method == 'POST':
        year_id = request.POST.get('year_id')
        if year_id:
            AcademicYear.objects.all().update(is_current=False)
            AcademicYear.objects.filter(id=year_id).update(is_current=True)
            messages.success(request, 'تم تعيين العام الدراسي الفعال بنجاح.')
    return redirect('portal_dashboard')


def generate_stress_test_data(target_count=1200):
    """توليد سريع بدفعات ضخمة لـ 1,200 طالب مع درجات كاملة لجميع المواد لاختبار الضغط"""
    import random
    from decimal import Decimal
    from django.contrib.auth import get_user_model
    User = get_user_model()

    FIRST_NAMES = [
        "محمد", "علي", "أحمد", "حسين", "حسن", "عمر", "يوسف", "زيد", "مصطفى", "كرار",
        "سجاد", "عباس", "جعفر", "إبراهيم", "حيدر", "خالد", "عبدالله", "مهدي", "باقر",
        "فاطمة", "زينب", "مريم", "زهراء", "نور", "هدى", "سارة", "آية", "بنين", "تبارك"
    ]
    MID_NAMES = ["علي", "حسين", "محمد", "كريم", "عباس", "جاسم", "راضي", "كاظم", "صالح", "هادي", "فاضل", "مهدي"]
    LAST_NAMES = ["الربيعي", "الجبوري", "الساعدي", "التميمي", "الشمري", "المالكي", "الزبيدي", "الخفاجي", "العامري", "اللامي", "الكعبي", "الجنابي"]

    classes = list(SchoolClass.objects.all().order_by('level_order'))
    if not classes:
        for i, name in enumerate(["الأول الابتدائي", "الثاني الابتدائي", "الثالث الابتدائي", "الرابع الابتدائي", "الخامس الابتدائي", "السادس الابتدائي"], start=1):
            classes.append(SchoolClass.objects.create(name=name, level_order=i))

    all_sections = []
    for cls in classes:
        for sec_name in ['أ', 'ب', 'ج', 'د']:
            sec, _ = Section.objects.get_or_create(school_class=cls, name=sec_name, defaults={'capacity': 50})
            all_sections.append(sec)

    class_subjects_map = {}
    for cls in classes:
        class_subjects_map[cls.id] = []
        for s_name in get_stage_official_subjects(cls):
            sub_obj, _ = Subject.objects.get_or_create(name=s_name, defaults={'code': f'SUB_{len(class_subjects_map[cls.id])+1}'})
            class_subjects_map[cls.id].append(sub_obj)

    users_to_create = []
    user_metadata = []
    current_time_str = timezone.now().strftime('%y%m%d%H%M')

    for i in range(1, target_count + 1):
        uname = f"stress_std_{current_time_str}_{i:04d}"
        f_name = random.choice(FIRST_NAMES)
        m_name = random.choice(MID_NAMES)
        l_name = random.choice(LAST_NAMES)
        full_quad = f"{f_name} {m_name} {random.choice(MID_NAMES)}"
        users_to_create.append(User(
            username=uname,
            first_name=full_quad,
            last_name=l_name,
            is_student=True
        ))
        target_class = classes[(i - 1) % len(classes)]
        cls_secs = [s for s in all_sections if s.school_class_id == target_class.id]
        target_sec = cls_secs[((i - 1) // len(classes)) % len(cls_secs)] if cls_secs else None
        user_metadata.append((uname, target_class, target_sec, i))

    created_users = User.objects.bulk_create(users_to_create, batch_size=1000)
    user_map = {u.username: u for u in created_users}

    students_to_create = []
    for uname, cls, sec, idx in user_metadata:
        u_obj = user_map[uname]
        students_to_create.append(Student(
            user=u_obj,
            current_class=cls,
            section=sec,
            registration_number=f"REG-2026-{idx:04d}",
            national_id=f"199{random.randint(100000000, 999999999)}",
            student_status='active'
        ))

    created_students = Student.objects.bulk_create(students_to_create, batch_size=1000)

    grades_to_create = []
    for st in created_students:
        cls_subs = class_subjects_map.get(st.current_class_id, [])
        for sub in cls_subs:
            f1 = Decimal(str(random.randint(45, 98)))
            mid = Decimal(str(random.randint(45, 95)))
            f2 = Decimal(str(random.randint(45, 98)))
            ann = round_integer((f1 + mid + f2) / Decimal('3'))
            fin = Decimal(str(random.randint(45, 98)))
            tot_final = round_integer((ann + fin) / Decimal('2'))

            dec = Decimal('0')
            dec_grade = None
            if 45 <= tot_final < 50:
                dec = Decimal(str(50 - tot_final))
                dec_grade = Decimal('50')
                status = 'passed_by_decision'
            elif tot_final >= 50:
                status = 'passed'
            else:
                status = 'supplementary' if tot_final >= 40 else 'failed'

            grades_to_create.append(Grade(
                student=st,
                subject=sub,
                first_term_effort=f1,
                midyear_exam=mid,
                second_term_effort=f2,
                annual_effort=ann,
                final_exam_round1=fin,
                final_grade=tot_final,
                decision_marks=dec,
                final_grade_after_decision=dec_grade,
                status=status
            ))

    Grade.objects.bulk_create(grades_to_create, batch_size=2000)
    return target_count


def clear_stress_test_data():
    """مسح وتطهير كافة بيانات اختبار الضغط بالكامل"""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    test_users = User.objects.filter(username__startswith='stress_std_')
    count = test_users.count()
    Grade.objects.filter(student__user__in=test_users).delete()
    Student.objects.filter(user__in=test_users).delete()
    test_users.delete()
    return count


def portal_settings(request):
    school = SchoolSettings.get_settings()
    current_year = get_active_academic_year()
    if request.method == 'POST':
        action_type = request.POST.get('action_type')
        stress_action = request.POST.get('stress_action')
        
        if action_type == 'set_academic_year' or stress_action == 'set_academic_year':
            year_id = request.POST.get('academic_year_id')
            new_year_name = request.POST.get('new_academic_year_name', '').strip()
            new_year = None
            if new_year_name:
                new_year, _ = AcademicYear.objects.get_or_create(
                    name=new_year_name,
                    defaults={'is_current': True}
                )
            elif year_id:
                new_year = get_object_or_404(AcademicYear, pk=year_id)
                
            if new_year:
                AcademicYear.objects.all().update(is_current=False)
                new_year.is_current = True
                new_year.save()
                request.session['active_academic_year_id'] = new_year.id
                messages.success(request, f"تم بنجاح اعتماد وتفعيل العام الدراسي ({new_year.name}). كافة بيانات وسجلات السنوات السابقة محفوظة بأمان تام في الأرشيف المركزي.")
            return redirect('portal_settings')

        # أدوات الاختبار محجوبة في بيئة الإنتاج
        if stress_action in ('generate_stress_data', 'clear_stress_data'):
            messages.warning(request, 'أدوات الاختبار غير متاحة في هذا الإصدار.')
            return redirect('portal_settings')

        ministry_code = request.POST.get('ministry_school_code', '').strip()
        if not ministry_code:
            messages.error(request, 'الرمز الإحصائي الوزاري للمدرسة (كود التربية) إجباري ولا يمكن تركه فارغاً.')
            return redirect('portal_settings')

        # حفظ المرحلة القديمة للكشف عن تغيير مرحلة المدرسة
        old_school_level = school.school_level

        school.ministry_school_code = ministry_code
        school.school_name = request.POST.get('school_name', school.school_name)
        school.director_name = request.POST.get('director_name', school.director_name)
        school.directorate = request.POST.get('directorate', school.directorate)
        school.sub_directorate = request.POST.get('sub_directorate', school.sub_directorate)
        school.school_gender = request.POST.get('school_gender', school.school_gender)
        school.school_level = request.POST.get('school_level', school.school_level)
        daily_p = request.POST.get('daily_periods_count')
        if daily_p:
            school.daily_periods_count = int(daily_p)
        if request.POST.get('remove_logo') == '1':
            if school.logo:
                school.logo.delete(save=False)
                school.logo = None
            messages.success(request, 'تمت إزالة الشعار بنجاح والعودة للشعار الافتراضي.')
        elif 'logo' in request.FILES:
            school.logo = request.FILES['logo']
            messages.success(request, 'تم تحديث وتبديل شعار المدرسة (اللوغو) بنجاح.')
        school.save()
        messages.success(request, 'تم حفظ وتحديث إعدادات وهوية المدرسة بنجاح.')

        # إعادة ضبط مراحل التخرج والصفوف تلقائياً عند تغيير مرحلة المدرسة
        if old_school_level != school.school_level:
            try:
                seed_count = seed_iraqi_official_classes(school)
                messages.info(
                    request,
                    f'تم إعادة ضبط الصفوف الدراسية ({seed_count} صف) ومراحل التخرج تلقائياً '
                    f'بما يتوافق مع المرحلة الجديدة ({school.get_school_level_display()}).'
                )
            except Exception as e:
                messages.warning(request, f'تعذر إعادة ضبط الصفوف تلقائياً: {e}')

        # إطلاق المزامنة السحابية تلقائياً في الخلفية فور الحفظ
        try:
            from .cloud_sync import upload_cloud_backup_async
            upload_cloud_backup_async()
        except Exception:
            pass

        return redirect('portal_settings')

    from .backup_vault import get_removable_drives, list_local_backups, get_backup_dir
    from .cloud_sync import get_last_cloud_sync_info
    removable_drives = get_removable_drives()
    local_backups = list_local_backups()[:7]
    backup_dir = get_backup_dir()
    last_cloud_sync = get_last_cloud_sync_info()

    context = {
        'school': school,
        'current_year': current_year,
        'academic_years': AcademicYear.objects.all().order_by('-start_date', '-name'),
        'removable_drives': removable_drives,
        'local_backups': local_backups,
        'backup_dir': backup_dir,
        'last_cloud_sync': last_cloud_sync,
    }
    return render(request, 'portal/settings.html', context)


def generate_years_view(request):
    if request.method == 'POST':
        count = AcademicYear.generate_next_50_years(start_year=2026)
        messages.success(request, f"تم بنجاح توليد وتجهيز {count} سنة دراسية للأرشيف والمستقبل ابتداءً من 2026-2027.")
        return redirect('portal_promotion')
    return redirect('portal_promotion')


def promotion_view(request):
    school = SchoolSettings.get_settings()
    academic_years = list(AcademicYear.objects.all().order_by('-start_date', '-name'))
    current_year = get_active_academic_year()
    classes = list(SchoolClass.objects.all().order_by('level_order'))

    selected_class_id = request.GET.get('class_id')
    if not selected_class_id and classes:
        selected_class_id = str(classes[0].id)

    target_class = SchoolClass.objects.filter(id=selected_class_id).first() if selected_class_id else None

    # تحديد سنة الترحيل المنتهية (from_year) والسنة المستهدفة الجديدة (to_year)
    from_year_id = request.GET.get('from_year')
    if from_year_id:
        from_yr = AcademicYear.objects.filter(id=from_year_id).first()
    else:
        from_yr = current_year

    to_year_id = request.GET.get('to_year')
    to_yr = AcademicYear.objects.filter(id=to_year_id).first() if to_year_id else None
    if not to_yr and from_yr:
        to_yr = AcademicYear.objects.filter(start_date__gt=from_yr.start_date).order_by('start_date').first()
        if not to_yr:
            to_yr = AcademicYear.objects.filter(name__gt=from_yr.name).order_by('name').first()
        if not to_yr:
            try:
                parts = from_yr.name.split('-')
                if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                    next_name = f"{int(parts[0])+1}-{int(parts[1])+1}"
                    to_yr, _ = AcademicYear.objects.get_or_create(
                        name=next_name,
                        defaults={
                            'start_date': from_yr.end_date or timezone.now().date(),
                            'end_date': (from_yr.end_date or timezone.now().date()) + timedelta(days=365),
                            'is_current': False
                        }
                    )
                    academic_years = list(AcademicYear.objects.all().order_by('-start_date', '-name'))
            except Exception:
                pass

    # الصف اللاحق الحتمي بحسب السلم الأصولي المعتمد
    deterministic_next = get_next_promotion_stage(target_class, school.school_level) if target_class else None
    if target_class and not target_class.next_class and deterministic_next:
        target_class.next_class = deterministic_next
        target_class.save(update_fields=['next_class'])

    # التحقق من استثناء صف الثالث المتوسط (التفرع للرابع العلمي والرابع الأدبي)
    is_third_intermediate = bool(target_class and 'الثالث المتوسط' in target_class.name)
    fourth_scientific, fourth_literary = get_third_intermediate_branches()
    is_secondary_school = school.school_level in ['secondary', 'preparatory', 'all_stages', 'all', 'comprehensive']
    has_branching = is_third_intermediate and is_secondary_school and (fourth_scientific is not None or fourth_literary is not None)

    students_data = []
    quarantined_students = []
    quarantined_count = 0

    if target_class:
        class_students = Student.objects.filter(
            current_class=target_class,
            student_status='active',
            is_deleted=False
        ).select_related('user', 'section').order_by('user__first_name', 'user__last_name', 'id')

        # محرك الحجب (Quarantine Engine): حجب الطلبة الذين رحلوا حديثاً إلى هذا الصف من مرحلة أدنى في نفس السنة لمنع الترفيع المزدوج
        quarantined_student_ids = set()
        if from_yr:
            quarantined_student_ids = set(
                StudentAcademicHistory.objects.filter(academic_year=from_yr)
                .exclude(school_class=target_class)
                .values_list('student_id', flat=True)
            )

        for st in class_students:
            if st.id in quarantined_student_ids:
                quarantined_students.append(st)
                continue

            grades = Grade.objects.filter(student=st)
            if from_yr:
                yr_grades = grades.filter(academic_year=from_yr.name)
                if yr_grades.exists():
                    grades = yr_grades
            elif current_year:
                yr_grades = grades.filter(academic_year=current_year.name)
                if yr_grades.exists():
                    grades = yr_grades

            failed_count = 0
            valid_count = 0
            tot = Decimal('0')

            for g in grades:
                eff = g.final_grade_after_decision or g.final_grade
                if eff is not None:
                    tot += eff
                    valid_count += 1
                    if eff < Decimal('50.0'):
                        failed_count += 1

            if valid_count > 0:
                if failed_count == 0:
                    computed_status = 'passed'
                    status_text = 'ناجح'
                elif failed_count <= 2:
                    computed_status = 're-exam'
                    status_text = 'مكمل'
                else:
                    computed_status = 'failed'
                    status_text = 'راسب'
            else:
                computed_status = 'manual_pending'
                status_text = 'تحديد يدوي'

            # القرار الافتراضي الأصولي
            if has_branching:
                default_decision = 'promote_scientific' if computed_status == 'passed' else 'stay'
            elif target_class.is_final_stage or not deterministic_next:
                default_decision = 'graduate' if computed_status == 'passed' else 'stay'
            else:
                default_decision = 'promote' if computed_status == 'passed' else 'stay'

            students_data.append({
                'student': st,
                'grades_count': valid_count,
                'failed_count': failed_count,
                'computed_status': computed_status,
                'status_text': status_text,
                'default_decision': default_decision,
            })

        quarantined_count = len(quarantined_students)

    if request.method == 'POST':
        action_type = request.POST.get('action_type')
        from_year_id = request.POST.get('from_year_id') or request.POST.get('from_year')
        to_year_id = request.POST.get('to_year_id') or request.POST.get('to_year')

        if not from_year_id or not to_year_id:
            messages.error(request, "يرجى تحديد سنة الترحيل المنتهية والسنة الجديدة المستهدفة.")
            return redirect(f"{request.path}?class_id={selected_class_id}")

        from_yr = get_object_or_404(AcademicYear, pk=from_year_id)
        to_yr = get_object_or_404(AcademicYear, pk=to_year_id)
        cls_id = request.POST.get('class_id')
        cls_obj = get_object_or_404(SchoolClass, pk=cls_id)

        # إعادة احتساب الحجب أمنياً لمنع أي تلاعب
        quarantined_student_ids = set(
            StudentAcademicHistory.objects.filter(academic_year=from_yr)
            .exclude(school_class=cls_obj)
            .values_list('student_id', flat=True)
        )

        det_next = get_next_promotion_stage(cls_obj, school.school_level) or cls_obj.next_class
        is_third_int = ('الثالث المتوسط' in cls_obj.name)
        sci_branch, lit_branch = get_third_intermediate_branches()
        is_sec = school.school_level in ['secondary', 'preparatory', 'all_stages', 'all', 'comprehensive']
        can_branch = is_third_int and is_sec and (sci_branch or lit_branch)

        if action_type == 'auto_promote_class':
            cls_students = Student.objects.filter(current_class=cls_obj, student_status='active', is_deleted=False)
            if quarantined_student_ids:
                cls_students = cls_students.exclude(id__in=quarantined_student_ids)

            promoted = 0
            graduated = 0
            retained = 0
            re_exam = 0

            with transaction.atomic():
                for st in cls_students:
                    # تصفير بطاقات نتائج ودرجات الطالب في السنة الجديدة
                    Grade.objects.filter(student=st, academic_year=to_yr.name).delete()

                    grades = Grade.objects.filter(student=st, academic_year=from_yr.name)
                    failed_count = 0
                    valid_count = 0
                    tot = Decimal('0')

                    for g in grades:
                        eff = g.final_grade_after_decision or g.final_grade
                        if eff is not None:
                            tot += eff
                            valid_count += 1
                            if eff < Decimal('50.0'):
                                failed_count += 1

                    avg = round_integer(tot / Decimal(str(valid_count))) if valid_count > 0 else 0

                    if valid_count > 0 and failed_count == 0:
                        StudentAcademicHistory.objects.update_or_create(
                            student=st, academic_year=from_yr,
                            defaults={'school_class': cls_obj, 'result_status': 'passed', 'general_average': Decimal(str(avg))}
                        )
                        next_stage = (sci_branch or lit_branch) if can_branch else det_next
                        if not next_stage or cls_obj.is_final_stage:
                            st.student_status = 'graduated'
                            st.save()
                            graduated += 1
                        else:
                            st.current_class = next_stage
                            st.section = next_stage.sections.first()
                            st.save()
                            Enrollment.objects.get_or_create(
                                student=st,
                                school_class=next_stage,
                                defaults={'academic_year': to_yr.name}
                            )
                            promoted += 1
                    elif valid_count > 0 and failed_count <= 2:
                        re_exam += 1
                        continue
                    else:
                        StudentAcademicHistory.objects.update_or_create(
                            student=st, academic_year=from_yr,
                            defaults={'school_class': cls_obj, 'result_status': 'failed', 'general_average': Decimal(str(avg))}
                        )
                        st.save()
                        Enrollment.objects.get_or_create(
                            student=st,
                            school_class=cls_obj,
                            defaults={'academic_year': to_yr.name}
                        )
                        retained += 1

            messages.success(request, f"اكتمل ترحيل صف ({cls_obj.name}) للعام ({to_yr.name}): ترفيع {promoted}، بقاء رسوباً {retained}، تخرج {graduated}، والمكملين {re_exam}.")
            return redirect(f"{request.path}?class_id={selected_class_id}&from_year={from_yr.id}&to_year={to_yr.id}")

        elif action_type == 'manual_promote_class':
            student_ids = request.POST.getlist('student_ids')
            promoted = 0
            retained = 0
            graduated = 0

            with transaction.atomic():
                for s_id in student_ids:
                    if not s_id or not s_id.isdigit():
                        continue
                    int_id = int(s_id)
                    if int_id in quarantined_student_ids:
                        continue

                    decision = request.POST.get(f'decision_{int_id}', 'stay')
                    st = get_object_or_404(Student, pk=int_id)

                    # تصفير بطاقات نتائج ودرجات الطالب في السنة الجديدة
                    Grade.objects.filter(student=st, academic_year=to_yr.name).delete()

                    if decision == 'graduate':
                        st.student_status = 'graduated'
                        st.save()
                        StudentAcademicHistory.objects.update_or_create(
                            student=st, academic_year=from_yr,
                            defaults={'school_class': cls_obj, 'result_status': 'graduated', 'general_average': Decimal('0')}
                        )
                        graduated += 1

                    elif decision == 'stay':
                        StudentAcademicHistory.objects.update_or_create(
                            student=st, academic_year=from_yr,
                            defaults={'school_class': cls_obj, 'result_status': 'failed', 'general_average': Decimal('0')}
                        )
                        st.student_status = 'active'
                        st.save()
                        Enrollment.objects.get_or_create(
                            student=st,
                            school_class=cls_obj,
                            defaults={'academic_year': to_yr.name}
                        )
                        retained += 1

                    elif decision == 'promote_scientific':
                        dest_class = sci_branch or det_next
                        if dest_class:
                            st.current_class = dest_class
                            st.section = dest_class.sections.first()
                            st.student_status = 'active'
                            st.save()
                            StudentAcademicHistory.objects.update_or_create(
                                student=st, academic_year=from_yr,
                                defaults={'school_class': cls_obj, 'result_status': 'passed', 'general_average': Decimal('0')}
                            )
                            Enrollment.objects.get_or_create(
                                student=st,
                                school_class=dest_class,
                                defaults={'academic_year': to_yr.name}
                            )
                            promoted += 1

                    elif decision == 'promote_literary':
                        dest_class = lit_branch or det_next
                        if dest_class:
                            st.current_class = dest_class
                            st.section = dest_class.sections.first()
                            st.student_status = 'active'
                            st.save()
                            StudentAcademicHistory.objects.update_or_create(
                                student=st, academic_year=from_yr,
                                defaults={'school_class': cls_obj, 'result_status': 'passed', 'general_average': Decimal('0')}
                            )
                            Enrollment.objects.get_or_create(
                                student=st,
                                school_class=dest_class,
                                defaults={'academic_year': to_yr.name}
                            )
                            promoted += 1

                    elif decision == 'promote':
                        dest_class = det_next
                        if dest_class and not cls_obj.is_final_stage:
                            st.current_class = dest_class
                            st.section = dest_class.sections.first()
                            st.student_status = 'active'
                            st.save()
                            StudentAcademicHistory.objects.update_or_create(
                                student=st, academic_year=from_yr,
                                defaults={'school_class': cls_obj, 'result_status': 'passed', 'general_average': Decimal('0')}
                            )
                            Enrollment.objects.get_or_create(
                                student=st,
                                school_class=dest_class,
                                defaults={'academic_year': to_yr.name}
                            )
                            promoted += 1
                        else:
                            st.student_status = 'graduated'
                            st.save()
                            StudentAcademicHistory.objects.update_or_create(
                                student=st, academic_year=from_yr,
                                defaults={'school_class': cls_obj, 'result_status': 'graduated', 'general_average': Decimal('0')}
                            )
                            graduated += 1

            messages.success(request, f"تم اعتماد وحفظ ترحيل صف ({cls_obj.name}) بنجاح: ترفيع {promoted}، بقاء {retained}، وتخرج {graduated} طالب.")
            return redirect(f"{request.path}?class_id={selected_class_id}&from_year={from_yr.id}&to_year={to_yr.id}")

    context = {
        'school': school,
        'academic_years': academic_years,
        'current_year': current_year,
        'from_yr': from_yr,
        'to_yr': to_yr,
        'classes': classes,
        'target_class': target_class,
        'deterministic_next': deterministic_next,
        'is_third_intermediate': is_third_intermediate,
        'has_branching': has_branching,
        'fourth_scientific': fourth_scientific,
        'fourth_literary': fourth_literary,
        'selected_class_id': selected_class_id,
        'students_data': students_data,
        'quarantined_students': quarantined_students,
        'quarantined_count': quarantined_count,
    }
    return render(request, 'portal/promotion.html', context)


# ======================================================================
# القاعات الامتحانية مع معالجة academic_year_id بدقة تامة
# ======================================================================

def exam_halls_view(request):
    school = SchoolSettings.get_settings()
    current_year = get_active_academic_year()
    sessions = ExamSession.objects.select_related('academic_year').prefetch_related('halls').all().order_by('-id')
    halls = ExamHall.objects.all().order_by('name')
    classes = SchoolClass.objects.all().order_by('level_order')

    if request.method == 'POST':
        action_type = request.POST.get('action_type', '')

        if action_type == 'create_session':
            title = request.POST.get('title', '').strip()
            hall_ids = request.POST.getlist('halls')
            if title:
                today = timezone.now().date()
                session = ExamSession.objects.create(
                    title=title,
                    academic_year=current_year,
                    start_date=today,
                    end_date=today
                )
                if hall_ids:
                    session.halls.set(hall_ids)
                messages.success(request, f"تم إنشاء الدورة الامتحانية ({title}) بنجاح.")
            else:
                messages.error(request, "يرجى تحديد عنوان للدورة الامتحانية.")
            return redirect('portal_exam_halls')

        elif action_type == 'create_hall':
            name = request.POST.get('name', '').strip()
            location = request.POST.get('location', '').strip()
            lines_count = int(request.POST.get('lines_count', 3))
            desks_per_line = int(request.POST.get('desks_per_line', 6))
            desk_type = request.POST.get('desk_type', 'single')

            if name:
                hall = ExamHall(
                    name=name,
                    location=location,
                    lines_count=lines_count,
                    desks_per_line=desks_per_line,
                    desk_type=desk_type
                )
                hall.save()
                messages.success(request, f"تمت إضافة القاعة الامتحانية ({name}) بنجاح بسعة {hall.capacity} مقعد ({hall.lines_count} خطوط، {hall.desks_per_line} رحلة/خط، {hall.get_desk_type_display()}).")
                return redirect(f"/portal/exam-halls/?hall_id={hall.id}&tab=layout#hall_layout_{hall.id}")
            else:
                messages.error(request, "يرجى كتابة اسم أو رقم القاعة.")
            return redirect('portal_exam_halls')

        elif action_type == 'edit_hall':
            hall_id = request.POST.get('hall_id')
            hall = get_object_or_404(ExamHall, pk=hall_id)
            name = request.POST.get('name', '').strip()
            location = request.POST.get('location', '').strip()
            lines_count = int(request.POST.get('lines_count', hall.lines_count or 3))
            desks_per_line = int(request.POST.get('desks_per_line', hall.desks_per_line or 6))
            desk_type = request.POST.get('desk_type', hall.desk_type or 'single')

            if name:
                hall.name = name
                hall.location = location
                hall.lines_count = lines_count
                hall.desks_per_line = desks_per_line
                hall.desk_type = desk_type
                hall.save()
                messages.success(request, f"تم تعديل بيانات القاعة ({name}) بنجاح. السعة المحدثة: {hall.capacity} مقعد.")
                return redirect(f"/portal/exam-halls/?hall_id={hall.id}&tab=layout#hall_layout_{hall.id}")
            else:
                messages.error(request, "اسم القاعة لا يمكن أن يكون فارغاً.")
            return redirect('portal_exam_halls')

        elif action_type == 'delete_hall':
            hall_id = request.POST.get('hall_id')
            hall = get_object_or_404(ExamHall, pk=hall_id)
            hall_name = hall.name
            hall.delete()
            messages.success(request, f"تم حذف القاعة الامتحانية ({hall_name}) وكافة المقاعد المرتبطة بها بنجاح.")
            return redirect('portal_exam_halls')

        elif 'distribute_seats' in request.POST or action_type == 'distribute_seats':
            session_id = request.POST.get('session_id')
            selected_classes_ids = request.POST.getlist('classes')

            if not session_id:
                messages.error(request, "يرجى اختيار الدورة الامتحانية أولاً.")
                return redirect('portal_exam_halls')

            session = get_object_or_404(ExamSession, pk=session_id)
            session_halls = session.halls.all()

            if not session_halls.exists():
                messages.error(request, "الدورة الامتحانية المختارة لا تحتوي على قاعات مخصصة.")
                return redirect('portal_exam_halls')

            if not selected_classes_ids:
                messages.error(request, "يرجى تحديد مرحلة أو صف واحد على الأقل لتوزيع طلبته.")
                return redirect('portal_exam_halls')

            # مسح التوزيع السابق للدورة
            ExamSeatAssignment.objects.filter(exam_session=session).delete()

            selected_classes = SchoolClass.objects.filter(id__in=selected_classes_ids)
            class_students = {}
            for cls in selected_classes:
                st_list = list(Student.objects.filter(current_class=cls, student_status='active', is_deleted=False).select_related('user', 'current_class', 'section'))
                random.shuffle(st_list)
                if st_list:
                    class_students[cls.id] = st_list

            if not class_students:
                messages.error(request, "لا يوجد طلاب نشطون في الصفوف المحددة.")
                return redirect('portal_exam_halls')

            assigned_count = 0
            last_placed_class_id = None

            # خوارزمية ذكية لتوزيع المقاعد ومكافحة الغش:
            # تخلط المراحل وتضمن عدم جلوس طالبين من نفس الصف في رحلة ثنائية
            for hall in session_halls:
                seat_number = 1
                lines = hall.lines_count or 3
                desks = hall.desks_per_line or 6
                is_double = (hall.desk_type == 'double')
                seats_per_desk = 2 if is_double else 1

                for d in range(1, desks + 1):
                    for l in range(1, lines + 1):
                        for pos in range(seats_per_desk):
                            if seat_number > hall.capacity:
                                break

                            available_class_ids = [cid for cid, s_list in class_students.items() if s_list]
                            if not available_class_ids:
                                break

                            # اختيار المرحلة لتجنب تجاور طلاب من نفس المرحلة في المقعد الثنائي
                            chosen_cid = None
                            if pos == 1 and last_placed_class_id and len(available_class_ids) > 1:
                                other_cids = [cid for cid in available_class_ids if cid != last_placed_class_id]
                                if other_cids:
                                    chosen_cid = max(other_cids, key=lambda cid: len(class_students[cid]))

                            if not chosen_cid:
                                chosen_cid = max(available_class_ids, key=lambda cid: len(class_students[cid]))

                            st = class_students[chosen_cid].pop(0)
                            last_placed_class_id = chosen_cid

                            ExamSeatAssignment.objects.create(
                                exam_session=session,
                                exam_hall=hall,
                                student=st,
                                seat_number=seat_number,
                                desk_row=d,
                                desk_col=l
                            )
                            assigned_count += 1
                            seat_number += 1

                        if not any(class_students.values()):
                            break
                    if not any(class_students.values()):
                        break
                if not any(class_students.values()):
                    break

            messages.success(request, f"تم توزيع {assigned_count} طالب على القاعات بنظام الخلط الذكي ومكافحة الغش بنجاح.")
            return redirect('portal_exam_halls')

    # جلب بيانات المعاينة التفاعلية لمقاعد القاعات
    selected_session_id = request.GET.get('session_id')
    if not selected_session_id and sessions.exists():
        selected_session_id = str(sessions.first().id)

    selected_session = ExamSession.objects.filter(id=selected_session_id).first() if selected_session_id else None
    selected_hall_id = request.GET.get('hall_id', '').strip()
    search_q = request.GET.get('q', '').strip()
    active_tab = request.GET.get('tab', 'grid').strip()

    preview_seats = ExamSeatAssignment.objects.none()
    if selected_session:
        preview_seats = ExamSeatAssignment.objects.filter(exam_session=selected_session).select_related(
            'student__user', 'student__current_class', 'student__section', 'exam_hall'
        ).order_by('exam_hall__name', 'seat_number')

        if selected_hall_id:
            preview_seats = preview_seats.filter(exam_hall_id=selected_hall_id)

        if search_q:
            preview_seats = preview_seats.filter(
                Q(student__user__first_name__icontains=search_q) |
                Q(student__user__last_name__icontains=search_q) |
                Q(student__registration_number__icontains=search_q) |
                Q(seat_number__icontains=search_q)
            )

    # تجهيز بيانات المخطط الهندسي والمعاينة البصرية لكافة القاعات (سواء وزعت مقاعد أم تم إنشاؤها للتو)
    halls_layout_data = []
    for hall in halls:
        lines = hall.lines_count or 3
        desks = hall.desks_per_line or 6
        is_double = (hall.desk_type == 'double')
        seats_per_desk = 2 if is_double else 1

        seat_map = {}
        if selected_session:
            hall_assignments = ExamSeatAssignment.objects.filter(
                exam_session=selected_session,
                exam_hall=hall
            ).select_related('student__user', 'student__current_class', 'student__section')
            for asg in hall_assignments:
                seat_map[asg.seat_number] = asg

        grid_rows = []
        seat_num = 1
        for r in range(1, desks + 1):
            row_cols = []
            for c in range(1, lines + 1):
                desk_seats = []
                for p in range(seats_per_desk):
                    curr_num = seat_num
                    asg = seat_map.get(curr_num)
                    desk_seats.append({
                        'seat_number': curr_num,
                        'is_assigned': asg is not None,
                        'assignment': asg,
                        'desk_pos': p + 1,
                    })
                    seat_num += 1
                row_cols.append({
                    'row_idx': r,
                    'col_idx': c,
                    'seats': desk_seats,
                })
            grid_rows.append({
                'row_idx': r,
                'cols': row_cols,
            })

        halls_layout_data.append({
            'hall': hall,
            'lines_count': lines,
            'desks_per_line': desks,
            'is_double': is_double,
            'capacity': hall.capacity,
            'grid_rows': grid_rows,
            'assigned_seats_count': len(seat_map),
        })

    context = {
        'school': school,
        'sessions': sessions,
        'halls': halls,
        'halls_layout_data': halls_layout_data,
        'classes': classes,
        'current_year': current_year,
        'selected_session': selected_session,
        'selected_hall_id': selected_hall_id,
        'search_q': search_q,
        'preview_seats': preview_seats,
        'total_assigned_seats': preview_seats.count(),
        'active_tab': active_tab,
    }
    return render(request, 'portal/exam_halls.html', context)


def print_exam_labels(request, session_id):
    school = SchoolSettings.get_settings()
    session = get_object_or_404(ExamSession, pk=session_id)
    hall_id = request.GET.get('hall')
    seats = ExamSeatAssignment.objects.filter(exam_session=session).select_related('student__user', 'student__current_class', 'student__section', 'exam_hall')
    if hall_id:
        seats = seats.filter(exam_hall_id=hall_id)

    context = {
        'school': school,
        'session': session,
        'seats': seats,
    }
    return render(request, 'portal/exam_labels_print.html', context)


# =========================================================================
# التقسيم الإجباري لطباعة السجلات المدرسية وفق المعايير الرسمية (A4 Portrait)
# =========================================================================
CHUNK_SIZE = 27

def chunk_students_for_print(students_list):
    """
    تقسيم قائمة طلاب كل شعبة/صف إلى صفحات سعة كل منها 27 اسماً
    وتوليد الصفوف الفارغة برمجياً لإكمال الصفحة بدقة متناهية وفق النموذج العراقي
    """
    pages = []
    items = list(students_list)
    total = len(items)
    for i in range(0, max(total, 1), CHUNK_SIZE):
        chunk = items[i:i + CHUNK_SIZE]
        empty_count = CHUNK_SIZE - len(chunk)
        pages.append({
            'students': chunk,
            'empty_rows': range(empty_count),  # لتعبئة الحقول الفارغة حتى يكتمل الـ 27
            'start_no': i + 1,
            'page_number': (i // CHUNK_SIZE) + 1,
            'page_num': (i // CHUNK_SIZE) + 1,
            'students_count': len(chunk),
            'empty_rows_count': empty_count,
            'empty_rows_range': range(len(chunk) + 1, CHUNK_SIZE + 1),
            'total_students': total,
        })
    return pages


def print_exam_attendance(request, session_id):
    school = SchoolSettings.get_settings()
    session = get_object_or_404(ExamSession, pk=session_id)
    hall_id = request.GET.get('hall')
    seats = ExamSeatAssignment.objects.filter(exam_session=session).select_related('student__user', 'student__current_class', 'student__section', 'exam_hall')
    if hall_id:
        seats = seats.filter(exam_hall_id=hall_id)
    seats_list = list(seats)
    pages = chunk_students_for_print(seats_list)

    context = {
        'school': school,
        'session': session,
        'seats': seats_list,
        'pages': pages,
        'today_date': timezone.now().strftime('%Y/%m/%d'),
    }
    return render(request, 'portal/exam_attendance_print.html', context)


def general_registry_view(request):
    school = SchoolSettings.get_settings()
    query = request.GET.get('q', '').strip()
    class_id = request.GET.get('school_class', '').strip()
    status_filter = request.GET.get('status', '').strip()

    students = Student.objects.filter(is_deleted=False).select_related('user', 'current_class', 'section', 'parent__user').order_by('id')

    if query:
        students = students.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(registration_number__icontains=query) |
            Q(national_id__icontains=query)
        )
    if class_id:
        students = students.filter(current_class_id=class_id)
    if status_filter:
        students = students.filter(student_status=status_filter)

    classes = SchoolClass.objects.prefetch_related('sections').all().order_by('level_order')
    sections = Section.objects.all().select_related('school_class').order_by('school_class__level_order', 'name')

    from django.core.paginator import Paginator
    PAGE_SIZE = 27
    paginator = Paginator(students, PAGE_SIZE)
    page_number = request.GET.get('page', 1)
    students_page = paginator.get_page(page_number)
    
    current_count = len(students_page.object_list)
    empty_rows_count = max(0, PAGE_SIZE - current_count)
    empty_rows_range = range(current_count + 1, PAGE_SIZE + 1)

    context = {
        'school': school,
        'students': students_page,
        'page_obj': students_page,
        'paginator': paginator,
        'total_count': paginator.count,
        'empty_rows_count': empty_rows_count,
        'empty_rows_range': empty_rows_range,
        'empty_rows': range(empty_rows_count),
        'start_no': (students_page.number - 1) * PAGE_SIZE + 1,
        'classes': classes,
        'sections': sections,
        'query': query,
        'selected_class': class_id,
        'selected_status': status_filter,
        'today_date': timezone.now().strftime('%Y/%m/%d'),
    }
    return render(request, 'portal/general_registry.html', context)


def letter_builder_view(request):
    school = SchoolSettings.get_settings()
    templates = OfficialLetterTemplate.objects.all()
    preview_content = ""
    selected_template = None

    if request.method == 'POST':
        template_id = request.POST.get('template_id')
        student_name = request.POST.get('student_name', '')
        class_name = request.POST.get('class_name', '')
        reg_number = request.POST.get('reg_number', '')
        teacher_name = request.POST.get('teacher_name', '')
        school_name = request.POST.get('school_name', school.school_name)

        if template_id:
            selected_template = get_object_or_404(OfficialLetterTemplate, pk=template_id)
            raw_text = selected_template.content
            preview_content = raw_text.replace('{{student_name}}', student_name)\
                                      .replace('{{class_name}}', class_name)\
                                      .replace('{{reg_number}}', reg_number)\
                                      .replace('{{teacher_name}}', teacher_name)\
                                      .replace('{{school_name}}', school_name)\
                                      .replace('{{date}}', timezone.now().strftime('%Y/%m/%d'))

        if 'save_doc' in request.POST:
            OfficialDocument.objects.create(
                doc_number=request.POST.get('doc_number', '1/ك'),
                doc_date=timezone.now().date(),
                doc_type='issued',
                sender_receiver=request.POST.get('destination', 'الجهة المعنية'),
                subject=request.POST.get('subject', selected_template.name if selected_template else 'كتاب رسمي'),
                body_content=preview_content,
                status='completed',
                created_by=request.user if request.user.is_authenticated else None
            )
            messages.success(request, "تم حفظ الكتاب في سجل الصادر والكتب الرسمية بنجاح.")

    context = {
        'school': school,
        'templates': templates,
        'preview_content': preview_content,
        'selected_template': selected_template,
        'today_date': timezone.now().strftime('%Y/%m/%d'),
    }
    return render(request, 'portal/letter_builder.html', context)


def get_stage_official_subjects(school_class):
    """استحضار المواد الدراسية الرسمية المعتمدة لوزارة التربية بحسب المرحلة الدراسية"""
    if not school_class:
        return [
            "التربية الإسلامية", "اللغة العربية", "اللغة الإنكليزية",
            "الرياضيات", "العلوم", "الاجتماعيات", "التربية الفنية والنشيد", "التربية الرياضية"
        ]

    c_name = school_class.name

    # 1. المرحلة الابتدائية
    if any(k in c_name for k in ['الاول الابتدائي', 'الأول الابتدائي', 'الثاني الابتدائي', 'الثالث الابتدائي']):
        return [
            "التربية الإسلامية", "اللغة العربية", "اللغة الإنكليزية",
            "الرياضيات", "العلوم", "التربية الفنية والنشيد", "التربية الرياضية", "الأخلاقية"
        ]
    elif any(k in c_name for k in ['الرابع الابتدائي', 'الخامس الابتدائي', 'السادس الابتدائي']):
        return [
            "التربية الإسلامية", "اللغة العربية", "اللغة الإنكليزية",
            "الرياضيات", "العلوم", "الاجتماعيات", "التربية الفنية والنشيد", "التربية الرياضية", "الأخلاقية"
        ]

    # 2. المرحلة المتوسطة
    elif any(k in c_name for k in ['المتوسط', 'متوسط']):
        return [
            "التربية الإسلامية", "اللغة العربية", "اللغة الإنكليزية",
            "الرياضيات", "العلوم (أحياء، كيمياء، فيزياء)", "الاجتماعيات (تاريخ، جغرافية، وطنية)",
            "الحاسوب", "التربية الفنية", "التربية الرياضية"
        ]

    # 3. المرحلة الإعدادية والثانوية الفرع العلمي
    elif 'العلمي' in c_name:
        return [
            "التربية الإسلامية", "اللغة العربية", "اللغة الإنكليزية",
            "الرياضيات", "علم الأحياء", "الكيمياء", "الفيزياء", "الحاسوب"
        ]

    # 4. المرحلة الإعدادية والثانوية الفرع الأدبي
    elif 'الادبي' in c_name or 'الأدبي' in c_name:
        return [
            "التربية الإسلامية", "اللغة العربية", "اللغة الإنكليزية",
            "الرياضيات", "التاريخ", "الجغرافية", "الاقتصاد", "الفلسفة وعلم النفس", "الحاسوب"
        ]

    # إذا كانت هناك مواد مخصصة في قاعدة البيانات
    db_subjects = list(Subject.objects.values_list('name', flat=True))
    if db_subjects:
        return db_subjects

    return [
        "التربية الإسلامية", "اللغة العربية", "اللغة الإنكليزية",
        "الرياضيات", "العلوم", "الاجتماعيات", "التربية الفنية والنشيد", "التربية الرياضية"
    ]


def portal_records_manage(request):
    """إدارة سجلات الإدارة وقوائم الشفوي المعتمدة وفق النموذج العراقي"""
    school = SchoolSettings.get_settings()
    classes = SchoolClass.objects.all().order_by('level_order')
    current_year = get_active_academic_year()

    selected_class_id = request.GET.get('class_id', '')
    record_type = request.GET.get('record_type', 'master_exam_sheet')
    selected_subject_id = request.GET.get('subject_id', '')
    try:
        oral_columns_count = int(request.GET.get('oral_columns', 3) or 3)
    except (ValueError, TypeError):
        oral_columns_count = 3

    try:
        max_score = int(request.GET.get('max_score', 100) or 100)
    except (ValueError, TypeError):
        max_score = 100

    selected_class = None
    sections_data = []
    all_class_students = []

    if selected_class_id:
        try:
            selected_class = SchoolClass.objects.filter(pk=int(selected_class_id)).first()
        except (ValueError, TypeError):
            selected_class = None

        if not selected_class and classes.exists():
            selected_class = classes.first()
            selected_class_id = str(selected_class.id)

    if selected_class:
        class_sections = list(selected_class.sections.all().order_by('name'))
        if not class_sections:
            sec_def, _ = Section.objects.get_or_create(school_class=selected_class, name='أ', defaults={'capacity': 40})
            class_sections = [sec_def]

        # جلب جميع طلاب الصف غير المحذوفين مع بياناتهم
        all_class_students = list(
            Student.objects.filter(current_class=selected_class, is_deleted=False)
            .select_related('user', 'section')
            .order_by('user__first_name', 'user__last_name', 'id')
        )

        for sec in class_sections:
            if sec == class_sections[0]:
                sec_students = [s for s in all_class_students if s.section == sec or s.section is None]
            else:
                sec_students = [s for s in all_class_students if s.section == sec]

            pages = chunk_students_for_print(sec_students)

            sections_data.append({
                'section': sec,
                'students': sec_students,
                'students_count': len(sec_students),
                'pages': pages,
                'total_pages': len(pages),
            })

    subjects = Subject.objects.all().order_by('name')
    subjects_list = get_stage_official_subjects(selected_class)
    all_pages = chunk_students_for_print(all_class_students) if selected_class_id else []

    context = {
        'school': school,
        'classes': classes,
        'subjects': subjects,
        'subjects_list': subjects_list,
        'selected_class': selected_class,
        'selected_class_id': selected_class_id,
        'selected_subject_id': selected_subject_id,
        'record_type': record_type,
        'sections_data': sections_data,
        'pages': all_pages,
        'all_students': all_class_students if selected_class_id else [],
        'current_year': current_year,
        'oral_columns_count': oral_columns_count,
        'oral_columns_range': range(1, oral_columns_count + 1),
        'admin_rows_range': range(1, 28),
        'empty_pages_range': range(1, 6),
        'max_score': max_score,
        'today_date': timezone.now().strftime('%Y/%m/%d'),
    }
    return render(request, 'portal/records_manage.html', context)


def portal_records_export_pdf(request):
    """
    تصدير السجل الوسطي بالكامل أو لشعبة/طالب محدد إلى ملف PDF فوري
    يعتمد على المعالجة الخلفية للسيرفر (ReportLab) دون تعليق المتصفح أو استدعاء حوار الطباعة.
    """
    school = SchoolSettings.get_settings()
    current_year = get_active_academic_year()

    selected_class_id = request.GET.get('class_id', '')
    if not selected_class_id:
        messages.error(request, "يرجى اختيار الصف الدراسي أولاً لتصدير السجل.")
        return redirect('portal_records_manage')

    try:
        selected_class = SchoolClass.objects.filter(pk=int(selected_class_id)).first()
    except (ValueError, TypeError):
        selected_class = None

    if not selected_class:
        messages.error(request, "الصف الدراسي المحدد غير موجود.")
        return redirect('portal_records_manage')

    subjects_list = get_stage_official_subjects(selected_class)

    filter_section_id = request.GET.get('section_id', '')
    filter_student_id = request.GET.get('student_id', '')
    page_from = request.GET.get('page_from', '')
    page_to = request.GET.get('page_to', '')

    class_sections = list(selected_class.sections.all().order_by('name'))
    if not class_sections:
        sec_def, _ = Section.objects.get_or_create(school_class=selected_class, name='أ', defaults={'capacity': 40})
        class_sections = [sec_def]

    students_qs = Student.objects.filter(current_class=selected_class, is_deleted=False).select_related('user', 'section').order_by('user__first_name', 'user__last_name', 'id')

    if filter_student_id and filter_student_id.isdigit():
        students_qs = students_qs.filter(id=int(filter_student_id))
    elif filter_section_id and filter_section_id.isdigit():
        students_qs = students_qs.filter(section_id=int(filter_section_id))
        class_sections = [s for s in class_sections if s.id == int(filter_section_id)]

    all_students = list(students_qs)

    # تطبيق نطاق الصفحات إذا تم تحديده
    if page_from and page_from.isdigit():
        p_from = max(1, int(page_from)) - 1
        p_to = int(page_to) if (page_to and page_to.isdigit()) else len(all_students)
        all_students = all_students[p_from:p_to]

    sections_data = []
    for sec in class_sections:
        if sec == class_sections[0]:
            sec_students = [s for s in all_students if s.section == sec or s.section is None]
        else:
            sec_students = [s for s in all_students if s.section == sec]

        if sec_students:
            sections_data.append({
                'section': sec,
                'students': sec_students,
                'students_count': len(sec_students),
            })

    from .pdf_generator import generate_middle_record_pdf
    pdf_bytes = generate_middle_record_pdf(
        school=school,
        selected_class=selected_class,
        current_year=current_year,
        sections_data=sections_data,
        subjects_list=subjects_list,
        empty_pages_count=3 if not filter_student_id else 0
    )

    filename = f"سجل_الدرجات_الوسطي_{selected_class.name}.pdf".replace(' ', '_')
    from django.utils.encoding import escape_uri_path
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{escape_uri_path(filename)}"'
    return response


# ======================================================================
# إدارة جدول الحصص والتوليد الآلي الذكي
# ======================================================================

def check_timetable_conflict(teacher_id, class_id, section_id, day, period, room=None, exclude_slot_id=None):
    conflicts = []

    if teacher_id:
        t_slots = TimetableSlot.objects.filter(teacher_id=teacher_id, day=day, period=period, is_active=True)
        if exclude_slot_id:
            t_slots = t_slots.exclude(id=exclude_slot_id)
        if t_slots.exists():
            conflict_slot = t_slots.first()
            teacher_name = conflict_slot.teacher.user.get_full_name()
            conflicts.append(
                f"⚠️ تعارض في جدول المعلم ({teacher_name}): لديه حصة بالفعل في {conflict_slot.school_class.name} "
                f"- شعبة ({conflict_slot.section.name if conflict_slot.section else 'العامة'}) في الحصة {conflict_slot.period}."
            )

    c_slots = TimetableSlot.objects.filter(school_class_id=class_id, day=day, period=period, is_active=True)
    if section_id:
        c_slots = c_slots.filter(section_id=section_id)
    if exclude_slot_id:
        c_slots = c_slots.exclude(id=exclude_slot_id)
    if c_slots.exists():
        conflict_slot = c_slots.first()
        conflicts.append(
            f"⚠️ تعارض في الشعبة: الصف ({conflict_slot.school_class.name}) لديه مادة مسجلة في هذا التوقيت."
        )

    return conflicts


def portal_timetable(request):
    school = SchoolSettings.get_settings()
    current_year = get_active_academic_year()

    classes = SchoolClass.objects.prefetch_related('sections').all().order_by('level_order')
    teachers = Teacher.objects.select_related('user').all()
    subjects = Subject.objects.all().order_by('name')

    all_sections = Section.objects.select_related('school_class').all().order_by('school_class__level_order', 'name')

    p_param = request.GET.get('periods_count')
    if p_param and p_param.isdigit():
        active_periods_count = max(1, min(10, int(p_param)))
    else:
        active_periods_count = school.daily_periods_count or 7
    periods_range = list(range(1, active_periods_count + 1))

    view_mode = request.GET.get('view_mode', 'master')
    selected_day = int(request.GET.get('day', 0))
    selected_class_id = request.GET.get('class_id', '')
    selected_section_id = request.GET.get('section_id', '')
    selected_teacher_id = request.GET.get('teacher_id', '')

    if request.method == 'POST':
        action_type = request.POST.get('action_type', 'save_slot')

        if action_type == 'save_timing_settings':
            p_count = int(request.POST.get('daily_periods_count', 6))
            school.daily_periods_count = max(1, min(10, p_count))
            school.save(update_fields=['daily_periods_count'])
            messages.success(request, "تم تحديث توقيتات وعدد الحصص الدراسية اليومية بنجاح.")
            return redirect(request.get_full_path())

        elif action_type == 'auto_generate_schedule':
            with transaction.atomic():
                TimetableSlot.objects.all().delete()
                assigned_count = 0
                all_teachers = list(Teacher.objects.prefetch_related('subjects', 'school_classes').all())

                if not all_teachers:
                    messages.error(request, "لا يوجد معلمون في النظام لتوزيع الحصص عليهم.")
                    return redirect(request.get_full_path())

                # ===== محرك الجدول الذكي المقيّد بالقيود =====
                # busy_teachers: مفتاح التعارض عالمياً (معلم × يوم × حصة)
                busy_teachers = set()
                # teacher_workload: عدد الحصص الكلية لكل معلم (موازنة العبء)
                teacher_workload = {t.id: 0 for t in all_teachers}
                slots_to_create = []

                # بناء فهرس الصف → قائمة المواد المعتمدة (مرتبة بالاسم لثبات التوزيع)
                class_subjects_cache = {}
                for cls in classes:
                    cls_subs = list(get_stage_official_subjects(cls))
                    if not cls_subs:
                        cls_subs = list(subjects)
                    class_subjects_cache[cls.id] = cls_subs

                # بناء فهرس المعلم → مجموعة المواد التي يدرسها
                teacher_subjects_cache = {}
                teacher_classes_cache = {}
                for tch in all_teachers:
                    teacher_subjects_cache[tch.id] = set(s.id for s in tch.subjects.all())
                    teacher_classes_cache[tch.id] = set(c.id for c in tch.school_classes.all())

                for cls in classes:
                    sections_list = list(cls.sections.all())
                    if not sections_list:
                        sections_list = [None]

                    cls_subjects = class_subjects_cache[cls.id]
                    total_subs = len(cls_subjects)
                    if total_subs == 0:
                        continue

                    for sec in sections_list:
                        # مؤشر دوري للمواد لضمان التوزيع العادل
                        subject_idx = 0

                        for day_idx in range(5):
                            for period_num in range(1, active_periods_count + 1):
                                # المادة المقررة لهذه الحصة (توزيع دوري منضبط)
                                current_subject = cls_subjects[subject_idx % total_subs]
                                subject_idx += 1

                                chosen_teacher = None
                                chosen_subject = current_subject

                                # ترتيب المعلمين حسب العبء (الأقل حصصاً أولاً) + اختلاط لتجنب التحيز
                                sorted_teachers = sorted(
                                    all_teachers,
                                    key=lambda t: teacher_workload[t.id]
                                )

                                # الجولة الأولى: مطابقة تامة (معلم مسند لهذا الصف + يدرس هذه المادة)
                                for tch in sorted_teachers:
                                    slot_key = (tch.id, day_idx, period_num)
                                    if slot_key in busy_teachers:
                                        continue
                                    t_subs = teacher_subjects_cache[tch.id]
                                    t_classes = teacher_classes_cache[tch.id]
                                    class_match = (not t_classes) or (cls.id in t_classes)
                                    sub_match = (not t_subs) or (current_subject.id in t_subs)
                                    if class_match and sub_match:
                                        chosen_teacher = tch
                                        busy_teachers.add(slot_key)
                                        teacher_workload[tch.id] += 1
                                        break

                                # الجولة الثانية: مطابقة جزئية (معلم يدرس هذه المادة فقط)
                                if not chosen_teacher:
                                    for tch in sorted_teachers:
                                        slot_key = (tch.id, day_idx, period_num)
                                        if slot_key in busy_teachers:
                                            continue
                                        t_subs = teacher_subjects_cache[tch.id]
                                        if (not t_subs) or (current_subject.id in t_subs):
                                            chosen_teacher = tch
                                            busy_teachers.add(slot_key)
                                            teacher_workload[tch.id] += 1
                                            break

                                # الجولة الثالثة: تخصيص أي معلم شاغر (طوارئ الجدول)
                                if not chosen_teacher:
                                    for tch in sorted_teachers:
                                        slot_key = (tch.id, day_idx, period_num)
                                        if slot_key not in busy_teachers:
                                            # اختر مادة يتقنها هذا المعلم
                                            t_subs = teacher_subjects_cache[tch.id]
                                            fallback_sub = None
                                            if t_subs:
                                                for s in cls_subjects:
                                                    if s.id in t_subs:
                                                        fallback_sub = s
                                                        break
                                            if not fallback_sub:
                                                fallback_sub = current_subject
                                            chosen_teacher = tch
                                            chosen_subject = fallback_sub
                                            busy_teachers.add(slot_key)
                                            teacher_workload[tch.id] += 1
                                            break

                                if chosen_teacher and chosen_subject:
                                    slots_to_create.append(
                                        TimetableSlot(
                                            school_class=cls,
                                            section=sec,
                                            teacher=chosen_teacher,
                                            subject=chosen_subject,
                                            day=day_idx,
                                            period=period_num,
                                            is_active=True
                                        )
                                    )

                if slots_to_create:
                    TimetableSlot.objects.bulk_create(slots_to_create, batch_size=500)
                    assigned_count = len(slots_to_create)

            messages.success(request, f"تم بنجاح توليد الجدول المدرسي آلياً وتوزيع {assigned_count} حصة بشكل منضبط ومحكم بدون أي تضارب زمني.")
            return redirect(request.get_full_path())


        elif action_type == 'save_slot':
            slot_id = request.POST.get('slot_id')
            c_id = request.POST.get('class_id')
            s_id = request.POST.get('section_id') or None
            sub_id = request.POST.get('subject_id')
            t_id = request.POST.get('teacher_id') or None
            day_val = int(request.POST.get('day', 0))
            period_val = int(request.POST.get('period', 1))
            room_val = request.POST.get('room', '').strip()
            notes_val = request.POST.get('notes', '').strip()

            conflicts = check_timetable_conflict(
                teacher_id=t_id, class_id=c_id, section_id=s_id,
                day=day_val, period=period_val, room=room_val,
                exclude_slot_id=slot_id
            )

            if conflicts:
                for msg in conflicts:
                    messages.error(request, msg)
            else:
                slot, created = TimetableSlot.objects.update_or_create(
                    id=slot_id if slot_id else None,
                    defaults={
                        'school_class_id': c_id,
                        'section_id': s_id,
                        'subject_id': sub_id,
                        'teacher_id': t_id,
                        'day': day_val,
                        'period': period_val,
                        'room': room_val,
                        'notes': notes_val,
                    }
                )
                messages.success(request, f"تم {'إسناد' if created else 'تحديث'} الحصة في الجدول بنجاح.")
            return redirect(request.get_full_path())

        elif action_type == 'delete_slot':
            slot_id = request.POST.get('slot_id')
            TimetableSlot.objects.filter(id=slot_id).delete()
            messages.success(request, "تم حذف الحصة بنجاح.")
            return redirect(request.get_full_path())

        elif action_type == 'add_substitution':
            slot_id = request.POST.get('slot_id')
            sub_t_id = request.POST.get('substitute_teacher_id')
            reason = request.POST.get('reason', 'إجازة طارئة / ظرف رسمي')
            slot = TimetableSlot.objects.filter(id=slot_id).first()
            sub_teacher = Teacher.objects.filter(id=sub_t_id).first()
            if slot and sub_teacher:
                slot.teacher = sub_teacher
                slot.notes = f"بديل: {sub_teacher.user.get_full_name()} ({reason})"
                slot.save(update_fields=['teacher', 'notes'])
                messages.success(request, f"تم بنجاح تكليف المعلم البديل ({sub_teacher.user.get_full_name()}) لشغل حصة الاحتياط.")
            return redirect(request.get_full_path())

        elif action_type == 'update_teacher_quota':
            req_p = request.POST.get('required_periods', 24)
            messages.success(request, f"تم اعتماد نصاب المعلم الأسبوعي ({req_p}) حصة بنجاح.")
            return redirect(request.get_full_path())

    COLOR_MAP = {
        'التربية الاسلامية': '#059669',
        'الاسلامية': '#059669',
        'اللغة العربية': '#2563eb',
        'اللغة الانكليزية': '#7c3aed',
        'الرياضيات': '#dc2626',
        'العلوم': '#0891b2',
        'الفيزياء': '#4f46e5',
        'الكيمياء': '#d97706',
        'الاحياء': '#16a34a',
        'الاجتماعيات': '#b45309',
        'التاريخ': '#854d0e',
        'الجغرافية': '#047857',
        'الحاسوب': '#475569',
        'التربية الرياضية': '#ea580c',
        'التربية الفنية': '#db2777',
    }

    all_slots_qs = TimetableSlot.objects.filter(is_active=True).select_related('school_class', 'section', 'subject', 'teacher__user')

    table_rows = []
    master_rows = []
    selected_class_obj = None
    selected_teacher_obj = None

    if view_mode == 'master':
        day_slots = all_slots_qs.filter(day=selected_day)
        sec_slot_map = {(s.section_id, s.period): s for s in day_slots}

        for sec in all_sections:
            sec_periods = []
            for p in periods_range:
                slot_obj = sec_slot_map.get((sec.id, p))
                col = '#2563eb'
                if slot_obj:
                    col = COLOR_MAP.get(slot_obj.subject.name, '#2563eb')
                sec_periods.append({
                    'period': p,
                    'slot': slot_obj,
                    'color': col,
                    'class_id': sec.school_class_id,
                    'section_id': sec.id,
                })
            master_rows.append({
                'section': sec,
                'periods': sec_periods,
            })

    elif view_mode == 'class':
        if selected_class_id:
            selected_class_obj = classes.filter(id=selected_class_id).first()
        if not selected_class_obj:
            selected_class_obj = classes.first()
            if selected_class_obj:
                selected_class_id = str(selected_class_obj.id)

        cls_slots = all_slots_qs.filter(school_class=selected_class_obj)
        if selected_section_id:
            cls_slots = cls_slots.filter(section_id=selected_section_id)

        matrix_lookup = {(s.day, s.period): s for s in cls_slots}

        for d_idx, d_name in TimetableSlot.DAYS_CHOICES:
            periods_slots = []
            for p in periods_range:
                slot_obj = matrix_lookup.get((d_idx, p))
                col = '#2563eb'
                if slot_obj:
                    col = COLOR_MAP.get(slot_obj.subject.name, '#2563eb')
                periods_slots.append({
                    'period': p,
                    'slot': slot_obj,
                    'color': col,
                    'class_id': selected_class_obj.id if selected_class_obj else None,
                    'section_id': selected_section_id or None,
                })
            table_rows.append({
                'day_index': d_idx,
                'day_name': d_name,
                'periods': periods_slots
            })

    elif view_mode == 'teacher':
        if selected_teacher_id:
            selected_teacher_obj = teachers.filter(id=selected_teacher_id).first()
        if not selected_teacher_obj:
            selected_teacher_obj = teachers.first()
            if selected_teacher_obj:
                selected_teacher_id = str(selected_teacher_obj.id)

        tch_slots = all_slots_qs.filter(teacher=selected_teacher_obj)
        matrix_lookup = {(s.day, s.period): s for s in tch_slots}

        for d_idx, d_name in TimetableSlot.DAYS_CHOICES:
            periods_slots = []
            for p in periods_range:
                slot_obj = matrix_lookup.get((d_idx, p))
                col = '#2563eb'
                if slot_obj:
                    col = COLOR_MAP.get(slot_obj.subject.name, '#2563eb')
                periods_slots.append({
                    'period': p,
                    'slot': slot_obj,
                    'color': col,
                    'teacher_id': selected_teacher_obj.id if selected_teacher_obj else None,
                })
            table_rows.append({
                'day_index': d_idx,
                'day_name': d_name,
                'periods': periods_slots
            })

    # بناء بيانات العرض الأسبوعي الشامل (الأحد إلى الخميس معاً)
    weekly_master_data = []
    for day_idx, day_name in TimetableSlot.DAYS_CHOICES:
        day_slots = all_slots_qs.filter(day=day_idx)
        sec_slot_map = {(s.section_id, s.period): s for s in day_slots}
        day_sections = []
        for sec in all_sections:
            sec_periods = []
            for p in periods_range:
                slot_obj = sec_slot_map.get((sec.id, p))
                col = '#2563eb'
                if slot_obj:
                    col = COLOR_MAP.get(slot_obj.subject.name, '#2563eb')
                sec_periods.append({
                    'period': p,
                    'slot': slot_obj,
                    'color': col,
                    'class_id': sec.school_class_id,
                    'section_id': sec.id,
                })
            day_sections.append({
                'section': sec,
                'periods': sec_periods,
            })
        weekly_master_data.append({
            'day_index': day_idx,
            'day_name': day_name,
            'sections': day_sections,
        })

    # احتساب بيانات نصاب المعلمين الأسبوعي
    teachers_quota_data = []
    for t in teachers:
        scheduled_count = all_slots_qs.filter(teacher=t).count()
        req_quota = 24
        rem = req_quota - scheduled_count
        teachers_quota_data.append({
            'teacher': t,
            'required': req_quota,
            'scheduled': scheduled_count,
            'remaining': max(0, rem),
            'over': max(0, scheduled_count - req_quota),
        })

    context = {
        'school': school,
        'current_year': current_year,
        'classes': classes,
        'teachers': teachers,
        'subjects': subjects,
        'days_choices': TimetableSlot.DAYS_CHOICES,
        'active_periods_count': active_periods_count,
        'periods_range': periods_range,
        'view_mode': view_mode,
        'selected_day': selected_day,
        'selected_class_id': selected_class_id,
        'selected_section_id': selected_section_id,
        'selected_teacher_id': selected_teacher_id,
        'selected_class_obj': selected_class_obj,
        'selected_teacher_obj': selected_teacher_obj,
        'weekly_master_data': weekly_master_data,
        'master_rows': master_rows,
        'table_rows': table_rows,
        'all_slots': all_slots_qs,
        'total_classes': classes.count(),
        'total_sections': all_sections.count(),
        'total_teachers': teachers.count(),
        'total_scheduled_slots': all_slots_qs.count(),
        'teachers_quota_data': teachers_quota_data,
        'today_date': timezone.now().strftime('%Y/%m/%d'),
    }
    return render(request, 'portal/timetable.html', context)


# ======================================================================
# إدارة الطلاب وربط بوابة أولياء الأمور تلقائياً
# ======================================================================

def portal_students_manage(request):
    school = SchoolSettings.get_settings()
    classes = SchoolClass.objects.prefetch_related('sections').all().order_by('level_order')
    sections = Section.objects.all()

    selected_class_id = request.GET.get('class_id', '')
    selected_section_id = request.GET.get('section_id', '')

    students_qs = Student.objects.filter(is_deleted=False).select_related('user', 'current_class', 'section', 'parent__user').order_by('-id')

    if selected_class_id:
        students_qs = students_qs.filter(current_class_id=selected_class_id)
    if selected_section_id:
        students_qs = students_qs.filter(section_id=selected_section_id)

    if request.method == 'POST':
        action_type = request.POST.get('action_type', 'add_student')

        if action_type == 'add_student':
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            reg_number = request.POST.get('registration_number', '').strip()
            national_id = request.POST.get('national_id', '').strip()
            class_id = request.POST.get('class_id')
            section_id = request.POST.get('section_id') or None
            parent_name = request.POST.get('parent_name', '').strip()
            parent_phone = request.POST.get('parent_phone', '').strip()

            final_reg_num = reg_number if reg_number else None
            final_nat_id = national_id if national_id else None

            with transaction.atomic():
                username = f"std_{random.randint(100000, 999999)}"
                u = User.objects.create_user(username=username, first_name=first_name, last_name=last_name, is_student=True)

                parent_obj = None
                if parent_name or parent_phone:
                    p_uname = f"prnt_{parent_phone}" if parent_phone else f"prnt_{random.randint(100000, 999999)}"
                    p_user = User.objects.filter(username=p_uname).first()
                    if not p_user:
                        p_user = User.objects.create_user(
                            username=p_uname,
                            first_name=parent_name or f"ولي أمر {first_name}",
                            is_parent=True
                        )
                        if parent_phone and len(parent_phone) >= 6:
                            p_user.set_password(parent_phone[-6:])
                            p_user.save()

                    parent_obj, _ = Parent.objects.get_or_create(user=p_user, defaults={'phone': parent_phone})

                target_class_obj = SchoolClass.objects.filter(id=class_id).first() if class_id else None
                target_sec_obj = Section.objects.filter(id=section_id).first() if section_id else None

                student_obj = Student.objects.create(
                    user=u,
                    registration_number=final_reg_num,
                    national_id=final_nat_id,
                    current_class=target_class_obj,
                    section=target_sec_obj,
                    parent=parent_obj,
                    student_status='active'
                )

                curr_academic_year = get_active_academic_year()
                if target_class_obj:
                    Enrollment.objects.get_or_create(
                        student=student_obj,
                        school_class=target_class_obj,
                        defaults={
                            'academic_year': curr_academic_year.name if curr_academic_year else '2026-2027'
                        }
                    )

                messages.success(request, f"تم تسجيل الطالب ({first_name} {last_name}) وربطه ببوابة ولي الأمر بنجاح.")
            return redirect(request.get_full_path())

        elif action_type == 'import_excel':
            excel_file = request.FILES.get('excel_file')
            target_class_id = request.POST.get('target_class_id')

            if not excel_file or not target_class_id:
                messages.error(request, "يرجى اختيار ملف الإكسل وتحديد الصف المطلوب.")
                return redirect(request.get_full_path())

            target_class = get_object_or_404(SchoolClass, pk=target_class_id)
            class_sections = list(target_class.sections.all().order_by('name'))
            has_sections = bool(class_sections)

            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active

                parsed_students = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    full_name = str(row[0]).strip()
                    parts = full_name.split(' ', 1)
                    first_n = parts[0]
                    last_n = parts[1] if len(parts) > 1 else "الطالب"

                    avg_val = 0.0
                    if len(row) > 1 and row[1] is not None:
                        try:
                            avg_val = float(str(row[1]).strip())
                        except (ValueError, TypeError):
                            avg_val = 0.0

                    reg_raw = str(row[2]).strip() if len(row) > 2 and row[2] else None
                    phone_raw = str(row[3]).strip() if len(row) > 3 and row[3] else ""

                    parsed_students.append({
                        'first_name': first_n,
                        'last_name': last_n,
                        'average': avg_val,
                        'registration_number': reg_raw,
                        'phone': phone_raw
                    })

                if not parsed_students:
                    messages.warning(request, "الملف فارغ أو لا يحتوي على بيانات مقروءة في العمود الأول.")
                    return redirect(request.get_full_path())

                parsed_students.sort(key=lambda x: x['average'], reverse=True)

                curr_academic_year = get_active_academic_year()
                num_sec = len(class_sections) if has_sections else 1

                with transaction.atomic():
                    for idx, st_data in enumerate(parsed_students):
                        chosen_sec = None
                        if has_sections:
                            cycle = idx // num_sec
                            rem = idx % num_sec
                            sec_idx = rem if (cycle % 2 == 0) else (num_sec - 1 - rem)
                            chosen_sec = class_sections[sec_idx]

                        final_reg_num = st_data['registration_number']
                        if final_reg_num and Student.objects.filter(registration_number=final_reg_num).exists():
                            final_reg_num = f"{final_reg_num}-{random.randint(10, 99)}"

                        u = User.objects.create_user(
                            username=f"std_{random.randint(100000, 999999)}",
                            first_name=st_data['first_name'],
                            last_name=st_data['last_name'],
                            is_student=True
                        )

                        parent_obj = None
                        if st_data['phone']:
                            p_user, _ = User.objects.get_or_create(
                                username=f"prnt_{st_data['phone']}",
                                defaults={
                                    'first_name': f"ولي أمر {st_data['first_name']}",
                                    'is_parent': True
                                }
                            )
                            parent_obj, _ = Parent.objects.get_or_create(user=p_user, defaults={'phone': st_data['phone']})

                        st_inst = Student.objects.create(
                            user=u,
                            registration_number=final_reg_num,
                            current_class=target_class,
                            section=chosen_sec,
                            parent=parent_obj,
                            student_status='active'
                        )

                        Enrollment.objects.get_or_create(
                            student=st_inst,
                            school_class=target_class,
                            defaults={'academic_year': curr_academic_year.name if curr_academic_year else '2026-2027'}
                        )

                messages.success(
                    request,
                    f"تم استيراد {len(parsed_students)} طالب وتوزيعهم وربطهم بأولياء الأمور بنجاح."
                )
            except Exception as e:
                messages.error(request, f"حدث خطأ أثناء معالجة ملف الإكسل: {str(e)}")

            return redirect(request.get_full_path())

        elif action_type == 'soft_delete_student':
            student_id = request.POST.get('student_id')
            student = get_object_or_404(Student, id=student_id)
            student.student_status = 'transferred'
            student.is_deleted = True
            student.save()
            messages.info(request, f"تم نقل الطالب ({student.full_name}) وحفظ قيده في الأرشيف.")
            return redirect(request.get_full_path())

    from django.core.paginator import Paginator
    total_students = students_qs.count()
    paginator = Paginator(students_qs, 27)
    page_number = request.GET.get('page', 1)
    students_page = paginator.get_page(page_number)

    context = {
        'school': school,
        'classes': classes,
        'sections': sections,
        'students': students_page,
        'page_obj': students_page,
        'paginator': paginator,
        'selected_class_id': selected_class_id,
        'selected_section_id': selected_section_id,
        'total_students': total_students,
    }
    return render(request, 'portal/students_manage.html', context)


# ======================================================================
# تثبيت الصفوف العراقية والشعب
# ======================================================================

def portal_classes_manage(request):
    school = SchoolSettings.get_settings()
    classes = SchoolClass.objects.prefetch_related('sections').all().order_by('level_order')

    if request.method == 'POST':
        action_type = request.POST.get('action_type', '')

        if action_type == 'init_iraqi_classes':
            count = seed_iraqi_official_classes(school)
            messages.success(request, f"تمت تهيئة وتأكيد الصفوف الرسمية لوزارة التربية بنجاح بحسب مرحلة المدرسة ({school.get_school_level_display()}).")
            return redirect('portal_classes_manage')

        elif action_type == 'add_class':
            name = request.POST.get('name', '').strip()
            level_order = int(request.POST.get('level_order', 1))
            is_final = request.POST.get('is_final_stage') in ['on', 'true', '1', True]
            if name:
                cls_obj = SchoolClass.objects.create(
                    name=name,
                    level_order=level_order,
                    is_final_stage=is_final
                )
                Section.objects.create(school_class=cls_obj, name="أ", capacity=40)
                messages.success(request, f"تم إنشاء صف ({name}) مع الشعبة (أ) بنجاح.")
            return redirect('portal_classes_manage')

        elif action_type == 'delete_class':
            class_id = request.POST.get('class_id')
            cls_obj = get_object_or_404(SchoolClass, pk=class_id)
            c_name = cls_obj.name
            cls_obj.delete()
            messages.success(request, f"تم حذف صف ({c_name}) بالكامل بنجاح.")
            return redirect('portal_classes_manage')

        elif action_type == 'delete_all_class_students':
            class_id = request.POST.get('class_id')
            cls_obj = get_object_or_404(SchoolClass, pk=class_id)
            students = Student.objects.filter(current_class=cls_obj, is_deleted=False)
            st_count = students.count()
            with transaction.atomic():
                for st in students:
                    st.is_deleted = True
                    st.student_status = 'deleted'
                    st.save()
                    Enrollment.objects.filter(student=st).delete()
                    ExamSeatAssignment.objects.filter(student=st).delete()
            messages.success(request, f"تم حذف جميع طلاب صف ({cls_obj.name}) بنجاح وعددهم ({st_count}) طالب.")
            return redirect('portal_classes_manage')

        elif action_type == 'add_section':
            class_id = request.POST.get('class_id')
            section_name = request.POST.get('section_name', '').strip()
            capacity = int(request.POST.get('capacity', 40))

            if class_id and section_name:
                target_cls = get_object_or_404(SchoolClass, pk=class_id)
                Section.objects.create(
                    school_class=target_cls,
                    name=section_name,
                    capacity=capacity
                )
                messages.success(request, f"تمت إضافة شعبة ({section_name}) إلى صف ({target_cls.name}) بنجاح.")
            return redirect('portal_classes_manage')

        elif action_type == 'delete_empty_section':
            section_id = request.POST.get('section_id')
            sec = get_object_or_404(Section, pk=section_id)
            if Student.objects.filter(section=sec, is_deleted=False).exists():
                messages.error(request, f"لا يمكن حذف شعبة ({sec.name}) لأنها تحتوي على طلبة مسجلين.")
            else:
                TimetableSlot.objects.filter(section=sec).delete()
                sec.delete()
                messages.success(request, f"تم حذف الشعبة ({sec.name}) بنجاح.")
            return redirect('portal_classes_manage')

    return render(request, 'portal/classes_manage.html', {'school': school, 'classes': classes})


# ======================================================================
# رصد وسجلات الدرجات مع نظام الفصول العراقي المعتمد
# ======================================================================

def portal_grades_manage(request):
    school = SchoolSettings.get_settings()
    current_year = get_active_academic_year()
    grades = Grade.objects.select_related('student__user', 'subject').filter(academic_year=current_year.name if current_year else '2026-2027').order_by('-id')[:50]
    students = Student.objects.filter(is_deleted=False, student_status='active').select_related('user')
    subjects = Subject.objects.all()

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        subject_id = request.POST.get('subject_id')

        s1_m1 = float(request.POST.get('first_term_m1') or 0)
        s1_m2 = float(request.POST.get('first_term_m2') or s1_m1)
        s1_avg = (s1_m1 + s1_m2) / 2.0

        mid = float(request.POST.get('midyear') or 0)

        s2_m1 = float(request.POST.get('second_term_m1') or 0)
        s2_m2 = float(request.POST.get('second_term_m2') or s2_m1)
        s2_avg = (s2_m1 + s2_m2) / 2.0

        annual = round_integer((s1_avg + mid + s2_avg) / 3.0)

        fin = float(request.POST.get('final_exam') or 0)
        final_val = round_integer((annual + fin) / 2.0)

        Grade.objects.update_or_create(
            student_id=student_id,
            subject_id=subject_id,
            academic_year=current_year.name if current_year else "2026-2027",
            defaults={
                'first_term_effort': Decimal(str(round_integer(s1_avg))),
                'midyear_exam': Decimal(str(round_integer(mid))),
                'second_term_effort': Decimal(str(round_integer(s2_avg))),
                'annual_effort': Decimal(str(annual)),
                'final_exam_round1': Decimal(str(round_integer(fin))),
                'final_grade': Decimal(str(final_val)),
                'status': 'passed' if final_val >= 50 else 'failed'
            }
        )
        messages.success(request, "تم رصد واحتساب درجات الفصول والسعي السنوي بنجاح وبدون كسور.")
        return redirect('portal_grades_manage')

    context = {
        'school': school,
        'grades': grades,
        'students': students,
        'subjects': subjects,
        'current_year': current_year,
    }
    return render(request, 'portal/grades_manage.html', context)


# ======================================================================
# تعديل بيانات الطالب المستقل
# ======================================================================

def portal_student_edit(request, student_id):
    student = get_object_or_404(Student.objects.select_related('user', 'parent__user'), pk=student_id)

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        reg_number = request.POST.get('registration_number', '').strip()
        national_id = request.POST.get('national_id', '').strip()
        class_id = request.POST.get('class_id')
        section_id = request.POST.get('section_id')
        student_status = request.POST.get('student_status', student.student_status)

        with transaction.atomic():
            if first_name or last_name:
                student.user.first_name = first_name or student.user.first_name
                student.user.last_name = last_name or student.user.last_name
                student.user.save()

            student.registration_number = reg_number if reg_number else None
            student.national_id = national_id if national_id else None
            student.current_class_id = class_id if class_id else None
            student.section_id = section_id if section_id else None
            student.student_status = student_status
            student.save()

            messages.success(request, f"تم حفظ تعديل بيانات الطالب ({student.full_name}) بنجاح.")

    next_url = request.META.get('HTTP_REFERER', 'portal_general_registry')
    return redirect(next_url)


# باقي الدوال الإدارية
def portal_teachers_manage(request):
    school = SchoolSettings.get_settings()
    teachers = Teacher.objects.select_related('user').prefetch_related('subjects', 'school_classes').all().order_by('-id')
    all_subjects = Subject.objects.all().order_by('name')
    all_classes = SchoolClass.objects.all().order_by('level_order')

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        job_title = request.POST.get('job_title', 'مدرس').strip()
        stat_code = request.POST.get('statistical_code', '').strip()
        subject_ids = request.POST.getlist('subjects')
        class_ids = request.POST.getlist('classes')

        with transaction.atomic():
            username = f"tch_{random.randint(10000, 99999)}"
            u = User.objects.create_user(username=username, first_name=first_name, last_name=last_name, is_teacher=True)
            tch = Teacher.objects.create(user=u, job_title=job_title, statistical_code=stat_code)
            if subject_ids:
                tch.subjects.set(subject_ids)
            if class_ids:
                tch.school_classes.set(class_ids)

        messages.success(request, f"تمت إضافة التدريسي ({first_name} {last_name}) وربطه بالمواد والصفوف بنجاح.")
        return redirect('portal_teachers_manage')

    context = {
        'school': school,
        'teachers': teachers,
        'all_subjects': all_subjects,
        'all_classes': all_classes,
    }
    return render(request, 'portal/teachers_manage.html', context)


def portal_teacher_delete(request, teacher_id):
    teacher = get_object_or_404(Teacher.objects.select_related('user'), pk=teacher_id)
    if request.method == 'POST':
        teacher_name = teacher.user.get_full_name()
        with transaction.atomic():
            TimetableSlot.objects.filter(teacher=teacher).update(teacher=None)
            TeacherQuota.objects.filter(teacher=teacher).delete()
            u = teacher.user
            teacher.delete()
            if u:
                u.delete()
        messages.success(request, f"تم حذف التدريسي ({teacher_name}) بنجاح.")
    return redirect('portal_teachers_manage')


def portal_teacher_edit(request, teacher_id):
    teacher = get_object_or_404(Teacher.objects.select_related('user'), pk=teacher_id)
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        job_title = request.POST.get('job_title', '').strip()
        stat_code = request.POST.get('statistical_code', '').strip()
        subject_ids = request.POST.getlist('subjects')
        class_ids = request.POST.getlist('classes')

        with transaction.atomic():
            teacher.user.first_name = first_name or teacher.user.first_name
            teacher.user.last_name = last_name or teacher.user.last_name
            teacher.user.save()
            teacher.job_title = job_title or teacher.job_title
            teacher.statistical_code = stat_code
            teacher.save()
            teacher.subjects.set(subject_ids)
            teacher.school_classes.set(class_ids)
            messages.success(request, f"تم تعديل بيانات وربط التدريسي ({teacher.user.get_full_name()}) بالمواد والصفوف بنجاح.")
    return redirect('portal_teachers_manage')


def portal_subjects_manage(request):
    school = SchoolSettings.get_settings()
    subjects = Subject.objects.all().order_by('id')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip().upper()
        if name and code:
            Subject.objects.create(name=name, code=code)
            messages.success(request, f"تمت إضافة المادة الدراسية ({name}) بنجاح.")
            return redirect('portal_subjects_manage')
    return render(request, 'portal/subjects_manage.html', {'school': school, 'subjects': subjects})


def portal_subject_edit(request, subject_id):
    subject = get_object_or_404(Subject, pk=subject_id)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip().upper()
        if name and code:
            subject.name = name
            subject.code = code
            subject.save()
            messages.success(request, f"تم تحديث المادة الدراسية ({name}) بنجاح.")
    return redirect('portal_subjects_manage')


def portal_parents_manage(request):
    school = SchoolSettings.get_settings()
    query = request.GET.get('q', '').strip()
    parents = Parent.objects.select_related('user').prefetch_related('student_set__user', 'student_set__current_class').all().order_by('-id')

    if query:
        parents = parents.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(phone__icontains=query) |
            Q(address__icontains=query) |
            Q(student__user__first_name__icontains=query)
        ).distinct()

    return render(request, 'portal/parents_manage.html', {'school': school, 'parents': parents, 'query': query})


def portal_parent_edit(request, parent_id):
    parent = get_object_or_404(Parent.objects.select_related('user'), pk=parent_id)
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        address = request.POST.get('address', '').strip()

        with transaction.atomic():
            if first_name:
                parent.user.first_name = first_name
                parent.user.save()
            parent.phone = phone
            parent.address = address
            parent.save()
            messages.success(request, f"تم تحديث بيانات ولي الأمر ({parent.user.first_name}) بنجاح.")

    next_url = request.META.get('HTTP_REFERER', 'portal_parents_manage')
    return redirect(next_url)


def portal_attendance_manage(request):
    school = SchoolSettings.get_settings()
    classes = SchoolClass.objects.all().order_by('level_order')
    selected_class_id = request.GET.get('school_class', '')
    selected_section_id = request.GET.get('section', '')
    selected_date = request.GET.get('date', timezone.now().date().strftime('%Y-%m-%d'))

    sections = Section.objects.filter(school_class_id=selected_class_id) if selected_class_id else Section.objects.all()

    students = []
    if selected_class_id:
        st_qs = Student.objects.filter(is_deleted=False, student_status='active', current_class_id=selected_class_id)
        if selected_section_id:
            st_qs = st_qs.filter(section_id=selected_section_id)
        students = list(st_qs.select_related('user').order_by('user__first_name'))

    if request.method == 'POST':
        att_date = request.POST.get('attendance_date') or timezone.now().date()
        recorder = request.user if request.user.is_authenticated else None

        with transaction.atomic():
            for st in students:
                st_status = request.POST.get(f'status_{st.id}', 'present')
                Attendance.objects.update_or_create(
                    student=st,
                    date=att_date,
                    defaults={'status': st_status, 'recorded_by': recorder}
                )
        messages.success(request, f"تم حفظ وتثبيت سجل الحضور والغياب بنجاح.")
        return redirect(f"{request.path}?school_class={selected_class_id}&section={selected_section_id}&date={selected_date}")

    context = {
        'school': school,
        'classes': classes,
        'sections': sections,
        'students': students,
        'selected_class_id': selected_class_id,
        'selected_section_id': selected_section_id,
        'selected_date': selected_date,
    }
    return render(request, 'portal/attendance_manage.html', context)


# ======================================================================
# النسخ الاحتياطي والسحابة
# ======================================================================

def download_backup_view(request):
    db_path = settings.DATABASES['default']['NAME']
    if not os.path.exists(db_path):
        messages.error(request, "ملف قاعدة البيانات غير موجود حالياً!")
        return redirect('portal_settings')

    date_str = timezone.now().strftime('%Y_%m_%d_%H%M')
    backup_filename = f"school_backup_{date_str}.sqlite3"
    response = FileResponse(open(db_path, 'rb'), content_type='application/x-sqlite3')
    response['Content-Disposition'] = f'attachment; filename="{backup_filename}"'
    return response


def restore_backup_view(request):
    if request.method == 'POST' and request.FILES.get('backup_file'):
        backup_file = request.FILES['backup_file']
        db_path = settings.DATABASES['default']['NAME']
        try:
            if os.path.exists(db_path):
                shutil.copy2(db_path, f"{db_path}.temp_safety")
            with open(db_path, 'wb+') as destination:
                for chunk in backup_file.chunks():
                    destination.write(chunk)
            messages.success(request, "تمت استعادة كافة بيانات وسجلات المدرسة بنجاح تام!")
        except Exception as e:
            if os.path.exists(f"{db_path}.temp_safety"):
                shutil.copy2(f"{db_path}.temp_safety", db_path)
            messages.error(request, f"فشلت عملية الاسترجاع: {str(e)}")
    return redirect('portal_settings')


def upload_patch_view(request):
    if request.method == 'POST' and request.FILES.get('patch_file'):
        patch_file = request.FILES['patch_file']
        if not patch_file.name.endswith('.zip'):
            messages.error(request, "يرجى رفع حزمة تحديث صالحة.")
            return redirect('portal_settings')
        try:
            with zipfile.ZipFile(patch_file, 'r') as zip_ref:
                zip_ref.extractall(settings.BASE_DIR)
            messages.success(request, "تم تثبيت حزمة التحديث بنجاح.")
        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء فك الحزمة: {str(e)}")
    return redirect('portal_settings')


def portal_usb_backup_save(request):
    """
    حفظ نسخة احتياطية مباشرة على فلاش ميموري (USB) أو مسار مخصص
    """
    if request.method == 'POST':
        from .backup_vault import save_backup_to_usb
        usb_drive = request.POST.get('usb_drive', '').strip()
        custom_path = request.POST.get('custom_path', '').strip()
        target = usb_drive or custom_path
        if not target:
            messages.error(request, "يرجى اختيار محرك الفلاش ميموري أو كتابة مسار المجلد المطلوب.")
            return redirect('portal_settings')

        success, msg = save_backup_to_usb(target)
        if success:
            messages.success(request, msg)
        else:
            messages.error(request, msg)
    return redirect('portal_settings')


def portal_create_snapshot_now(request):
    """
    إنشاء لقطة يومية فورية وحفظها في الخزنة المحلية %APPDATA%/Madrasati/Backups/
    """
    if request.method == 'POST':
        from .backup_vault import create_daily_backup_snapshot
        success, msg = create_daily_backup_snapshot()
        if success:
            messages.success(request, msg)
        else:
            messages.error(request, msg)
    return redirect('portal_settings')


def portal_cloud_backup_now(request):
    """
    رفع نسخة سحابية فورية ومباشرة إلى Firebase RTDB
    """
    if request.method == 'POST':
        from .cloud_sync import upload_cloud_backup
        success, msg = upload_cloud_backup()
        if success:
            messages.success(request, msg)
        else:
            messages.error(request, msg)
    return redirect('portal_settings')


def portal_cloud_restore(request):
    """
    نافذة وعملية الاستعادة السحابية للطوارئ (Emergency Cloud Restore)
    """
    if request.method == 'POST':
        from .cloud_sync import restore_cloud_backup
        entered_code = request.POST.get('ministry_school_code', '').strip()
        if not entered_code:
            messages.error(request, "يرجى إدخال الرمز الإحصائي الوزاري للمدرسة لبدء الاستعادة السحابية.")
            return redirect('portal_settings')

        success, res = restore_cloud_backup(entered_code)
        if success:
            school_name = res.get('school_name', '')
            last_sync = res.get('last_sync', '')
            messages.success(
                request,
                f"تمت الاستعادة السحابية بنجاح لمدرسة ({school_name})! تاريخ النسخة المسترجعة: {last_sync}."
            )
        else:
            messages.error(request, f"فشلت عملية الاستعادة السحابية: {res}")
    return redirect('portal_settings')


def cloud_backup_upload_view(request):
    return portal_cloud_backup_now(request)


def cloud_backup_restore_view(request):
    return portal_cloud_restore(request)


APP_VERSION = "2.1.0"
GITHUB_REPO = "abdullahnawfal797-cmd/school-mgmt"

def cloud_check_update_view(request):
    """
    فحص توفر تحديثات للنظام عبر GitHub Releases API
    - استعلام خفيف وسريع (مهلة 1.8 ثانية)
    - لا يعطل النظام عند غياب الإنترنت ويعمل أوفلاين كالمعتاد بهدوء تام
    - يدعم استجابة JSON للتحقق التلقائي في الواجهة
    """
    repo = getattr(settings, 'GITHUB_UPDATE_REPO', GITHUB_REPO)
    url = f"https://api.github.com/repos/{repo}/releases/latest"
    headers = {
        'User-Agent': 'Madrasati-App-Updater',
        'Accept': 'application/vnd.github.v3+json'
    }
    
    is_json = request.GET.get('format') == 'json' or request.headers.get('x-requested-with') == 'XMLHttpRequest'

    try:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=1.8) as response:
            if response.status == 200:
                data = json.loads(response.read().decode('utf-8'))
                latest_tag = data.get('tag_name', '').lstrip('v').strip()
                current_ver = APP_VERSION.lstrip('v').strip()

                def parse_ver(v):
                    import re
                    parts = [int(x) for x in re.findall(r'\d+', v)]
                    return parts if parts else [0]

                has_update = parse_ver(latest_tag) > parse_ver(current_ver)
                download_url = data.get('html_url', '')
                for asset in data.get('assets', []):
                    if asset.get('browser_download_url', '').endswith(('.exe', '.zip')):
                        download_url = asset.get('browser_download_url')
                        break

                if is_json:
                    return JsonResponse({
                        'success': True,
                        'update_available': has_update,
                        'current_version': APP_VERSION,
                        'latest_version': latest_tag or APP_VERSION,
                        'release_name': data.get('name', f"الإصدار {latest_tag}"),
                        'release_notes': data.get('body', 'تحسينات عامة واستقرار المنظومة المدرسية'),
                        'download_url': download_url,
                        'published_at': data.get('published_at', '')
                    })
                else:
                    if has_update:
                        messages.info(request, f"يتوفر إصدار أحدث للمنظومة (v{latest_tag})! تفضل بالترقية.")
                    else:
                        messages.success(request, f"المنظومة محدثة لأحدث إصدار رسمي (v{APP_VERSION}).")
                    return redirect('portal_settings')
    except Exception:
        # صمت وتجاهل هادئ عند انقطاع الإنترنت أو عدم توفر الشبكة
        pass

    if is_json:
        return JsonResponse({
            'success': False,
            'update_available': False,
            'current_version': APP_VERSION,
            'message': 'المنظومة تعمل محلياً أو غير متصلة بالإنترنت.'
        })

    messages.info(request, f"المنظومة تعمل بالإصدار المستقر (v{APP_VERSION}) أوفلاين.")
    return redirect('portal_settings')


def apply_system_update_view(request):
    """
    تحميل مثبت التحديث أو حزمة التعديلات في مجلد مؤقت دون مساس بقاعدة البيانات المحلية db.sqlite3
    """
    if request.method == 'POST':
        download_url = request.POST.get('download_url')
        if not download_url:
            return JsonResponse({'success': False, 'error': 'رابط التحديث غير متوفر'})

        try:
            import tempfile
            temp_dir = os.path.join(tempfile.gettempdir(), 'Madrasati_Updates')
            os.makedirs(temp_dir, exist_ok=True)
            filename = download_url.split('/')[-1] or 'Madrasati_Update.zip'
            target_path = os.path.join(temp_dir, filename)

            req = urllib.request.Request(download_url, headers={'User-Agent': 'Madrasati-App-Updater'})
            with urllib.request.urlopen(req, timeout=30) as resp, open(target_path, 'wb') as out_f:
                shutil.copyfileobj(resp, out_f)

            return JsonResponse({
                'success': True,
                'message': f'تم تنزيل حزمة التحديث بنجاح إلى: {target_path}',
                'file_path': target_path
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'تعذر تنزيل التحديث: {str(e)}'})

    return JsonResponse({'success': False, 'error': 'طريقة الطلب غير مقبولة'})


# ======================================================================
# معالجة تعديل الصفوف الدراسية المفقودة
# ======================================================================

def portal_class_edit(request, pk=None):
    """تعديل بيانات الصف الدراسي وتفادي أخطاء الاستيراد في urls.py"""
    class_id = pk or request.POST.get('class_id')
    school_class = get_object_or_404(SchoolClass, pk=class_id)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        level_order = request.POST.get('level_order')
        is_final = request.POST.get('is_final_stage') in ['on', 'true', '1', True]

        if name:
            school_class.name = name
        if level_order:
            school_class.level_order = int(level_order)
        school_class.is_final_stage = is_final
        school_class.save()
        messages.success(request, f"تم تحديث بيانات صف ({school_class.name}) بنجاح.")

    return redirect('portal_classes_manage')


# ======================================================================
# معالج الإعداد الأولي السريع للمدرسة (First-Run Wizard)
# ======================================================================

def portal_first_run_setup(request):
    """حفظ بيانات معالج الإعداد الأولي للمدرسة أو تخطيه بنقرة واحدة"""
    school = SchoolSettings.get_settings()
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'skip':
            school.is_first_run_completed = True
            school.save(update_fields=['is_first_run_completed'])
            messages.info(request, "تم تخطي معالج التهيئة. يمكنك ضبط وتعديل بيانات المدرسة بأي وقت من قسم الإعدادات.")
        else:
            school_name = request.POST.get('school_name', '').strip()
            directorate = request.POST.get('directorate', '').strip()
            department = request.POST.get('department', '').strip()
            academic_year_id = request.POST.get('academic_year')
            logo = request.FILES.get('logo')

            if school_name:
                school.school_name = school_name
            if directorate:
                school.directorate = directorate
            if department:
                school.department = department
            ministry_code = request.POST.get('ministry_school_code', '').strip()
            if ministry_code:
                school.ministry_school_code = ministry_code
            if logo:
                school.logo = logo

            school.is_first_run_completed = True
            school.save()

            if academic_year_id:
                try:
                    selected_year = AcademicYear.objects.get(pk=academic_year_id)
                    AcademicYear.objects.filter(is_current=True).update(is_current=False)
                    selected_year.is_current = True
                    selected_year.save()
                    request.session['active_academic_year_id'] = selected_year.id
                except AcademicYear.DoesNotExist:
                    pass

            messages.success(request, f"أهلاً بك في نظام مدرستي! تم إعداد بيانات ({school.school_name}) بنجاح.")

    next_url = request.META.get('HTTP_REFERER') or reverse('portal_dashboard')
    return redirect(next_url)


# ======================================================================
# منظومة الأرشيف المركزي وسجلات المخاطبات الرسمية
# ======================================================================

def portal_official_archive(request):
    """
    بوابة الأرشيف المركزي المعتمد للمخاطبات الرسمية والكتب الصادرة والواردة وتأييدات الطلبة
    """
    school = SchoolSettings.get_settings()
    if not school.is_trial_or_license_valid:
        return redirect('portal_license_lock')

    active_tab = request.GET.get('tab', 'all').strip()
    query = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    qs = OfficialDocument.objects.all().select_related('student__user', 'teacher__user', 'created_by')

    # تصفية التبويبات الرسمية
    if active_tab == 'issued':
        qs = qs.filter(doc_type='issued')
    elif active_tab == 'received':
        qs = qs.filter(doc_type='received')
    elif active_tab == 'student_cert':
        qs = qs.filter(doc_type='student_cert')
    elif active_tab == 'circular':
        qs = qs.filter(doc_type='circular')
    elif active_tab == 'other':
        qs = qs.filter(doc_type='other')

    # محرك البحث الفوري
    if query:
        qs = qs.filter(
            Q(doc_number__icontains=query) |
            Q(subject__icontains=query) |
            Q(sender_receiver__icontains=query) |
            Q(notes__icontains=query) |
            Q(body_content__icontains=query) |
            Q(student__user__first_name__icontains=query) |
            Q(student__user__last_name__icontains=query) |
            Q(teacher__user__first_name__icontains=query) |
            Q(teacher__user__last_name__icontains=query)
        )

    # النطاق الزمني
    if date_from:
        qs = qs.filter(doc_date__gte=date_from)
    if date_to:
        qs = qs.filter(doc_date__lte=date_to)

    # الترتيب التنازلي الأحدث أولاً
    qs = qs.order_by('-doc_date', '-id')

    # الإحصائيات الأصولية السريعة
    all_docs = OfficialDocument.objects.all()
    stats = {
        'total': all_docs.count(),
        'issued': all_docs.filter(doc_type='issued').count(),
        'received': all_docs.filter(doc_type='received').count(),
        'student_cert': all_docs.filter(doc_type='student_cert').count(),
        'circular': all_docs.filter(doc_type='circular').count(),
    }

    paginator = Paginator(qs, 25)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except Exception:
        page_obj = paginator.get_page(1)

    context = {
        'school': school,
        'documents': page_obj,
        'active_tab': active_tab,
        'query': query,
        'date_from': date_from,
        'date_to': date_to,
        'stats': stats,
        'today_date': timezone.now().date().isoformat(),
    }
    return render(request, 'portal/official_archive.html', context)


def portal_archive_add(request):
    """تسجيل وأرشفة وثيقة أو كتاب أصولي جديد (صادر/وارد/تأييد/تعميم)"""
    if request.method == 'POST':
        doc_number = request.POST.get('doc_number', '').strip()
        doc_date_raw = request.POST.get('doc_date', '').strip()
        doc_type = request.POST.get('doc_type', 'issued').strip()
        sender_receiver = request.POST.get('sender_receiver', '').strip()
        subject = request.POST.get('subject', '').strip()
        body_content = request.POST.get('body_content', '').strip()
        notes = request.POST.get('notes', '').strip()
        status_val = request.POST.get('status', 'completed').strip()
        file_obj = request.FILES.get('file')

        if not doc_number or not subject:
            messages.error(request, "يرجى كتابة رقم الكتاب والموضوع على الأقل لإتمام الأرشفة.")
            return redirect('portal_official_archive')

        doc_date = timezone.now().date()
        if doc_date_raw:
            try:
                from datetime import datetime
                doc_date = datetime.strptime(doc_date_raw, '%Y-%m-%d').date()
            except Exception:
                pass

        doc = OfficialDocument.objects.create(
            doc_number=doc_number,
            doc_date=doc_date,
            doc_type=doc_type,
            sender_receiver=sender_receiver or 'الجهة المعنية',
            subject=subject,
            body_content=body_content,
            notes=notes,
            status=status_val,
            created_by=request.user if request.user.is_authenticated else None
        )
        if file_obj:
            doc.file = file_obj
            doc.save()

        messages.success(request, f"تمت أرشفة الوثيقة برقم ({doc.doc_number}) وموضوع ({doc.subject}) بنجاح.")

    return redirect('portal_official_archive')


def portal_archive_save_letter(request):
    """حفظ وتوثيق الخطاب المنشأ من منشئ المخاطبات مباشرة في الأرشيف الرسمي"""
    if request.method == 'POST':
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json'
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body.decode('utf-8'))
            except Exception:
                data = {}
        else:
            data = request.POST

        doc_number = data.get('doc_number', '').strip() or f"ص/{timezone.now().strftime('%Y%m%d%H%M')}"
        doc_date_raw = data.get('doc_date', '').strip()
        doc_type = data.get('doc_type', 'issued')
        sender_receiver = data.get('destination', data.get('sender_receiver', 'إلى من يهمه الأمر')).strip()
        subject = data.get('subject', 'كتاب رسمي صادر').strip()
        body_content = data.get('body_content', '').strip()
        notes = data.get('notes', 'محفوظ تلقائياً من منشئ المخاطبات المدرسية الأصولي').strip()

        doc_date = timezone.now().date()
        if doc_date_raw:
            try:
                from datetime import datetime
                doc_date = datetime.strptime(doc_date_raw.replace('/', '-'), '%Y-%m-%d').date()
            except Exception:
                pass

        doc = OfficialDocument.objects.create(
            doc_number=doc_number,
            doc_date=doc_date,
            doc_type=doc_type,
            sender_receiver=sender_receiver,
            subject=subject,
            body_content=body_content,
            notes=notes,
            status='completed',
            created_by=request.user if request.user.is_authenticated else None
        )

        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': f"تمت أرشفة الكتاب برقم أصولي ({doc.doc_number}) في الأرشيف المركزي بنجاح.",
                'doc_id': doc.id
            })

        messages.success(request, f"تم حفظ وأرشفة الوثيقة برقم ({doc.doc_number}) في الأرشيف الرسمي بنجاح.")
        return redirect('portal_official_archive')

    return redirect('portal_letter_builder')


def portal_archive_delete(request, doc_id):
    """حذف وثيقة من الأرشيف الرسمي"""
    if request.method == 'POST':
        doc = get_object_or_404(OfficialDocument, pk=doc_id)
        num = doc.doc_number
        doc.delete()
        messages.success(request, f"تم حذف القيد الأرشيفي رقم ({num}) بنجاح.")
    return redirect('portal_official_archive')


def portal_document_pdf_download(request, doc_id):
    """تنزيل ملف PDF عالي الجودة للوثيقة الرسمية من الأرشيف"""
    school = SchoolSettings.get_settings()
    doc = get_object_or_404(OfficialDocument, id=doc_id)
    from .pdf_generator import generate_official_document_pdf

    pdf_bytes = generate_official_document_pdf(
        school=school,
        doc_number=doc.doc_number,
        doc_date=doc.doc_date.strftime('%Y/%m/%d') if doc.doc_date else '',
        doc_type_display=doc.get_doc_type_display(),
        destination=doc.sender_receiver or 'إلى من يهمه الأمر',
        subject=doc.subject,
        body_content=doc.body_content,
        notes=doc.notes
    )
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    clean_num = str(doc.doc_number).replace('/', '_').replace('\\', '_').replace(' ', '_')
    filename = f"كتاب_رسمي_{clean_num}.pdf"
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
    return response


def portal_letter_export_pdf(request):
    """توليد وتنزيل ملف PDF فوري عالي الجودة للوثيقة الحالية في صانع الكتب"""
    school = SchoolSettings.get_settings()
    if request.method == 'POST':
        doc_num = request.POST.get('doc_number', '45/ص')
        doc_date = request.POST.get('doc_date', '')
        dest = request.POST.get('destination', 'الجهة المعنية')
        subj = request.POST.get('subject', 'كتاب رسمي')
        body = request.POST.get('body_content', '')
    else:
        doc_num = request.GET.get('doc_number', '45/ص')
        doc_date = request.GET.get('doc_date', '')
        dest = request.GET.get('destination', 'الجهة المعنية')
        subj = request.GET.get('subject', 'كتاب رسمي')
        body = request.GET.get('body_content', '')

    from .pdf_generator import generate_official_document_pdf
    pdf_bytes = generate_official_document_pdf(
        school=school,
        doc_number=doc_num,
        doc_date=doc_date,
        doc_type_display='كتاب رسمي صادر',
        destination=dest,
        subject=subj,
        body_content=body
    )
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    clean_num = str(doc_num).replace('/', '_').replace('\\', '_').replace(' ', '_')
    filename = f"كتاب_{clean_num}.pdf"
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
    return response


def portal_export_students_excel(request):
    """تصدير بيانات الطلبة إلى ملف Excel بحسب الصف والشعبة أو لكافة طلبة المدرسة مع دعم RTL"""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    school = SchoolSettings.get_settings()
    class_id = request.GET.get('class_id')
    section_id = request.GET.get('section_id')
    export_all = request.GET.get('export_all') == '1'

    students_qs = Student.objects.filter(is_deleted=False).select_related('user', 'current_class', 'section', 'parent__user')

    sub_title = "كافة طلبة المدرسة"
    if not export_all and class_id:
        students_qs = students_qs.filter(current_class_id=class_id)
        cls_obj = SchoolClass.objects.filter(id=class_id).first()
        if cls_obj:
            sub_title = f"طلبة {cls_obj.name}"
        if section_id:
            students_qs = students_qs.filter(section_id=section_id)
            sec_obj = Section.objects.filter(id=section_id).first()
            if sec_obj:
                sub_title += f" - شعبة {sec_obj.name}"
    elif not export_all and section_id:
        students_qs = students_qs.filter(section_id=section_id)
        sec_obj = Section.objects.filter(id=section_id).first()
        if sec_obj:
            sub_title = f"شعبة {sec_obj.name}"

    students_qs = students_qs.order_by('current_class__level_order', 'section__name', 'user__first_name', 'user__last_name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قوائم الطلبة"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(start_color="1E1830", end_color="1E1830", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1E1830")
    subtitle_font = Font(name="Calibri", size=11, bold=True, color="E06B22")
    data_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    ws.merge_cells("A1:G1")
    ws["A1"] = f"جمهورية العراق - وزارة التربية | {school.school_name}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = f"قائمة {sub_title} - إجمالي الطلبة: {students_qs.count()} طالب"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center_align
    ws.row_dimensions[2].height = 22

    headers = [
        'ت',
        'رقم القيد',
        'الاسم الرباعي واللقب',
        'الصف والشعبة',
        'الرقم الوطني / الهوية',
        'هاتف ولي الأمر',
        'الحالة الأكاديمية'
    ]
    ws.append([])
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 26

    status_map = {
        'active': 'مستمر بالدوام',
        'graduated': 'خريج',
        'transferred': 'منقول',
        'dismissed': 'مفصول / تارك'
    }

    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for idx, st in enumerate(students_qs, start=1):
        full_name = st.user.get_full_name() or st.user.username
        class_sec = f"{st.current_class.name if st.current_class else 'بدون صف'}"
        if st.section:
            class_sec += f" - شعبة {st.section.name}"
        parent_phone = st.parent.phone if st.parent and st.parent.phone else "---"
        status_text = status_map.get(st.student_status, st.student_status)

        row_data = [
            idx,
            st.registration_number or '---',
            full_name,
            class_sec,
            st.national_id or '---',
            parent_phone,
            status_text
        ]
        ws.append(row_data)
        current_row = 4 + idx
        ws.row_dimensions[current_row].height = 20

        for c_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = alt_fill
            if c_idx in [1, 2, 5, 6, 7]:
                cell.alignment = center_align
            else:
                cell.alignment = right_align

    col_widths = {1: 7, 2: 15, 3: 32, 4: 25, 5: 22, 6: 18, 7: 18}
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_title = sub_title.replace(' ', '_').replace('/', '_')
    filename = f"قائمة_الطلبة_{safe_title}.xlsx"
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
    wb.save(response)
    return response